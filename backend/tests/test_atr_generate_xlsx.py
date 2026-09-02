# backend/tests/test_atr_generate_xlsx.py
from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from app.services.atr_generate_xlsx import build_atr_xlsx
from tests._atr_fixtures import build_atr_workbook_bytes  # Phase A fixture builder


def _item(**kw):
    base = dict(pos=1, supplier_article_code="6060", part_number="VR11S 1010 016 000",
                part_number_norm="111010016000", part_name="CARPET EMERGENCY EXIT HATCH",
                drawing_number_issue="VR11S 1010-10/D", category="CARPET", qty=1,
                weight_kg=Decimal("0.413"), po_pos="050", match_status="matched", row_order=1)
    base.update(kw)
    return SimpleNamespace(**base)


def _delivery(**kw):
    base = dict(set_title="SET 6 BED CCRC", po_number="4501119979", msn="830",
                ba_auftrag="1024738", atr_number="ACM-A350CRC-ATR-4545-01",
                ac_programme="A350",
                qa_signer="Cordula Kesseler i.A.", max_guaranteed_weight_kg=Decimal("211"))
    base.update(kw)
    return SimpleNamespace(**base)


def test_build_writes_header_rows_and_total():
    items = [_item(), _item(pos=2, supplier_article_code="11395",
                            part_number="VR11S 1010 048 000", part_number_norm="111010048000",
                            part_name="CARPET BEDS FWD", drawing_number_issue="VR11S 1010-27/A",
                            weight_kg=Decimal("1.218"), po_pos="300", row_order=2)]
    out = build_atr_xlsx(build_atr_workbook_bytes(), _delivery(), items)
    wb = load_workbook(BytesIO(out))
    ws = [w for w in wb.worksheets if w.sheet_state == "visible"][0]
    assert ws["A11"].value == "SET 6 BED CCRC"
    # the two part rows are present somewhere in the table region
    pns = [ws.cell(r, 3).value for r in range(14, ws.max_row + 1)]
    assert "VR11S 1010 016 000" in pns and "VR11S 1010 048 000" in pns
    # a "Total weight" label exists with the summed value 1.631
    totals = [(r, ws.cell(r, 8).value) for r in range(14, ws.max_row + 1)
              if str(ws.cell(r, 6).value or "").lower().startswith("total")]
    assert totals and abs(float(totals[0][1]) - 1.631) < 0.001


def test_build_raises_when_template_has_no_totals():
    from io import BytesIO
    from openpyxl import Workbook
    import pytest
    wb = Workbook(); ws = wb.active; ws.title = "CCRC 6 BED"
    ws["A11"] = "SET 6 BED CCRC"
    for i, h in enumerate(["PO Pos","Article","Part Number / Index","Part Name",
                           "Serial","Drawing","Qty","Weight [kg]"], start=1):
        ws.cell(13, i, h)
    ws["A14"] = "CARPET"
    ws.cell(15, 3, "VR11S 1010 016 000")  # a part row, but NO "Total" label anywhere
    bio = BytesIO(); wb.save(bio)
    with pytest.raises(ValueError):
        build_atr_xlsx(bio.getvalue(), _delivery(), [_item()])


def test_unmatched_row_is_red():
    items = [_item(match_status="unmatched", drawing_number_issue=None,
                   weight_kg=None, part_name="UNKNOWN")]
    out = build_atr_xlsx(build_atr_workbook_bytes(), _delivery(), items)
    wb = load_workbook(BytesIO(out))
    ws = [w for w in wb.worksheets if w.sheet_state == "visible"][0]
    red_rows = [r for r in range(14, ws.max_row + 1)
                if ws.cell(r, 3).value == "VR11S 1010 016 000"
                and (ws.cell(r, 3).fill.fgColor.rgb or "").endswith("FF0000")]
    assert red_rows, "unmatched part row should carry a red fill"
