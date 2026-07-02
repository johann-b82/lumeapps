"""In-memory ATR workbook builder for tests — avoids committing binaries."""
from io import BytesIO

from openpyxl import Workbook

DEFAULT_PARTS = [
    # (article_code, part_number, part_name, serial, drawing, qty, weight)
    ("6060", "VR11S 1010 016 000", "CARPET EMERGENCY EXIT HATCH", "N/A", "VR11S 1010-10/D", 1, "0.413"),
    ("11395", "VR11S 1010 048 000", "CARPET BEDS FWD", "N/A", "VR11S 1010-27/A", 1, "1.218"),
]


def build_atr_workbook_bytes(parts=None, visible_title="CCRC 6 BED",
                             set_title="SET 6 BED CCRC") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = visible_title
    ws["A1"] = "Customer:"
    ws["D1"] = "Diehl Aviation Laupheim GmbH"
    ws["D2"] = "A350 XWB"
    ws["A3"] = "Work Package:"
    ws["D3"] = "Soft Furnishing for Flight and Cabin Crew Rest Compartments"
    ws["G3"] = 25
    ws["D4"] = "PTS 2552 0015 01, Issue 02"
    ws["G4"] = "C9312"
    ws["D5"] = "ACM-A350CRC-ATP-002 Issue 02"
    ws["D6"] = "ACM-A350CRC-SES-003 Issue 03"
    ws["D7"] = "PA-CO-BTS-2010-042-01-CRC_Soft Furnishing"
    ws["D8"] = "ACM GmbH - Woringer Strasse 11 - 87700 Memmingen"
    ws["G8"] = "N/A"
    ws["A11"] = set_title
    ws["F12"] = "Plattenwaage PW015"
    headers = ["PO Pos", "Supplier Article Code", "Part Number / Index",
               "Part Name", "Serial Number", "Drawing Number / Issue",
               "Qty", "Weight [kg]"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=13, column=i, value=h)
    ws["A14"] = "CARPET"
    rows = DEFAULT_PARTS if parts is None else parts
    r = 15
    for (art, pn, name, ser, draw, qty, wt) in rows:
        ws.cell(r, 2, art)
        ws.cell(r, 3, pn)
        ws.cell(r, 4, name)
        ws.cell(r, 5, ser)
        ws.cell(r, 6, draw)
        ws.cell(r, 7, qty)
        ws.cell(r, 8, wt)
        r += 1
    ws.cell(r + 1, 6, "Total weight")
    ws.cell(r + 1, 8, "1.631")
    wb.create_sheet("CCRC 8 BED").sheet_state = "hidden"
    wb.create_sheet("FCRC").sheet_state = "hidden"
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
