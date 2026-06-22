"""WM-2026 Tippspiel Excel parser.

Reads the "WM 2026 Tipps" sheet: one row per match with a tip per department.
Emits one normalised row per (match, department), with home/away team names
mapped onto the football-data feed names so the scoring service can match a
tip to its real result.

Quirks handled:
  * Group-header rows ("Gruppe A" with no fixture) are skipped.
  * Excel silently converted some "H:A" tips into times — ``time(3, 0)`` is
    the tip "3:0", ``time(1, 2)`` is "1:2". Recovered as (hour, minute).
  * Unknown team names produce an error entry (a mapping gap is visible).
"""
from __future__ import annotations

import io
import re
from datetime import date, datetime, time
from typing import Any

import openpyxl

from app.services.tippspiel_teams import to_feed_name

WM_YEAR = 2026
SHEET_NAME = "WM 2026 Tipps"
DEPT_COL_START = 3  # 0-based — columns D..H hold the five departments' tips
DEPT_COL_END = 8


def _parse_tip(val: Any) -> tuple[int, int] | None:
    if val is None:
        return None
    if isinstance(val, (time, datetime)):
        return (val.hour, val.minute)
    s = str(val).strip()
    if not s:
        return None
    m = re.match(r"^\s*(\d+)\s*:\s*(\d+)\s*$", s)
    if not m:
        return None
    return (int(m.group(1)), int(m.group(2)))


def _parse_date(val: Any) -> date | None:
    s = str(val or "").strip()
    m = re.search(r"(\d{1,2})\.(\d{1,2})", s)
    if not m:
        return None
    try:
        return date(WM_YEAR, int(m.group(2)), int(m.group(1)))
    except ValueError:
        return None


def parse_tippspiel_file(
    contents: bytes, filename: str
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[str]]:
    """Parse the Tipps workbook. Returns ``(rows, errors, departments)``."""
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(contents), data_only=True, read_only=True
        )
    except Exception as exc:  # pragma: no cover — malformed upload
        return [], [{"row": 0, "column": "", "message": f"unreadable: {exc}"}], []

    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.worksheets[0]
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return [], [{"row": 0, "column": "", "message": "empty sheet"}], []

    header = all_rows[0]
    departments = [
        str(header[i]).strip()
        for i in range(DEPT_COL_START, DEPT_COL_END)
        if i < len(header) and header[i]
    ]

    rows: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []

    for idx, r in enumerate(all_rows[1:], start=2):
        spiel = r[2] if len(r) > 2 else None
        if not spiel or "–" not in str(spiel):
            continue  # group header / blank row

        home_de, _, away_de = str(spiel).partition("–")
        home = to_feed_name(home_de)
        away = to_feed_name(away_de)
        if home is None or away is None:
            bad = home_de.strip() if home is None else away_de.strip()
            errors.append(
                {"row": idx, "column": "Spiel", "message": f"unknown team: {bad}"}
            )
            continue

        gruppe = str(r[0] or "").strip()
        mdate = _parse_date(r[1] if len(r) > 1 else None)

        for i, dept in enumerate(departments):
            col = DEPT_COL_START + i
            tip = _parse_tip(r[col] if len(r) > col else None)
            if tip is None:
                continue
            rows.append({
                "gruppe": gruppe,
                "home": home,
                "away": away,
                "match_date": mdate,
                "department": dept,
                "tip_home": tip[0],
                "tip_away": tip[1],
                "raw": {"spiel": str(spiel), "datum": str(r[1] if len(r) > 1 else "")},
            })

    return rows, errors, departments
