"""Parser für die Qualifikationsmatrizen (v1.90).

Die vier Bereichsdateien sind gleich aufgebaut, nur die Kopfzeile sitzt je
nach Datei auf Zeile 4 bis 6. Sie wird deshalb nicht fest verdrahtet, sondern
über den Text "Anzahl Mitarbeiter" in Spalte F gesucht.

Aufbau eines Blattes::

      B     C           D                F        G          I    J    K    L
    R6  | Nr | Kategorie | Bezeichnung    | Anzahl | Schnitt | <Name 1> | <Name 2>
    R7  | 1  | Allgemein | Deutsch        | 2      | 90      | AL | E% | AL | E%

Je Person zwei Spalten: Anforderungslevel (0-4) und Erfüllungsgrad (0-100 %).
"Anzahl Mitarbeiter" und "Durchschnitt" sind Excel-Formeln und werden bewusst
verworfen — beides ist aus den Bewertungen ableitbar.
"""
from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import date, datetime

from openpyxl import load_workbook

#: Erste Spalte mit Personendaten (I).
ERSTE_PERSONENSPALTE = 9
#: Spaltenindizes der festen Vorspalten.
SPALTE_NR = 2
SPALTE_KATEGORIE = 3
SPALTE_BEZEICHNUNG = 4
SPALTE_ANZAHL = 6
#: Blätter, die keine Matrix enthalten.
IGNORIERTE_BLAETTER = ("diagramm",)


@dataclass
class ParsedBewertung:
    person_index: int
    anforderungslevel: int | None
    erfuellungsgrad: int | None


@dataclass
class ParsedQualifikation:
    nr: int | None
    kategorie: str | None
    bezeichnung: str
    reihenfolge: int
    bewertungen: list[ParsedBewertung] = field(default_factory=list)


@dataclass
class ParsedMatrix:
    blatt: str
    titel: str | None
    stand: date | None
    personen: list[str]
    qualifikationen: list[ParsedQualifikation]


@dataclass
class ParsedDatei:
    dateiname: str
    matrizen: list[ParsedMatrix]
    warnungen: list[str] = field(default_factory=list)


def _text(wert: object) -> str | None:
    """Zellwert als aufgeräumter Text; leer wird zu None."""
    if wert is None:
        return None
    s = " ".join(str(wert).split())
    return s or None


def _ganzzahl(wert: object, unten: int, oben: int) -> int | None:
    """Zahl im erlaubten Bereich, sonst None ("N/A" kommt als Text vor)."""
    if isinstance(wert, bool) or not isinstance(wert, (int, float)):
        return None
    z = int(round(wert))
    return z if unten <= z <= oben else None


def _finde_kopfzeile(ws) -> int | None:
    """Zeile mit "Anzahl Mitarbeiter" — sitzt je nach Datei auf R4..R6."""
    for r in range(1, min(ws.max_row, 15) + 1):
        for c in range(1, 9):
            v = ws.cell(row=r, column=c).value
            if v and "Anzahl Mitarbeiter" in str(v):
                return r
    return None


def _stand(ws) -> date | None:
    """"Stand"-Datum aus der Kopfzeile (Beschriftung in F1, Wert daneben)."""
    for r in range(1, 4):
        for c in range(1, 10):
            v = ws.cell(row=r, column=c).value
            if v and str(v).strip().rstrip(":").lower() == "stand":
                nachbar = ws.cell(row=r, column=c + 1).value
                if isinstance(nachbar, datetime):
                    return nachbar.date()
                if isinstance(nachbar, date):
                    return nachbar
    return None


def _lies_blatt(ws) -> ParsedMatrix | None:
    kopf = _finde_kopfzeile(ws)
    if kopf is None:
        return None

    # Personen: ab Spalte I in Zweierschritten, Name steht über der AL-Spalte.
    personen: list[str] = []
    spalten: list[int] = []
    for c in range(ERSTE_PERSONENSPALTE, ws.max_column + 1, 2):
        name = _text(ws.cell(row=kopf, column=c).value)
        if name:
            personen.append(name)
            spalten.append(c)
    if not personen:
        return None

    qualifikationen: list[ParsedQualifikation] = []
    for r in range(kopf + 1, ws.max_row + 1):
        bezeichnung = _text(ws.cell(row=r, column=SPALTE_BEZEICHNUNG).value)
        if not bezeichnung:
            continue  # Leer- und Zwischenzeilen der Excel

        qual = ParsedQualifikation(
            nr=_ganzzahl(ws.cell(row=r, column=SPALTE_NR).value, 0, 10_000),
            kategorie=_text(ws.cell(row=r, column=SPALTE_KATEGORIE).value),
            bezeichnung=bezeichnung,
            reihenfolge=len(qualifikationen),
        )
        for i, c in enumerate(spalten):
            al = _ganzzahl(ws.cell(row=r, column=c).value, 0, 4)
            e = _ganzzahl(ws.cell(row=r, column=c + 1).value, 0, 100)
            if al is None and e is None:
                continue  # nichts eingetragen — keine Zelle anlegen
            qual.bewertungen.append(
                ParsedBewertung(person_index=i, anforderungslevel=al, erfuellungsgrad=e)
            )
        qualifikationen.append(qual)

    # Titel steht in der ersten Zeile ("Qualifikationsmatrix - Produktion").
    titel = None
    for c in range(1, 8):
        v = _text(ws.cell(row=1, column=c).value)
        if v and "matrix" in v.lower():
            titel = v
            break

    return ParsedMatrix(
        blatt=ws.title.strip(),
        titel=titel,
        stand=_stand(ws),
        personen=personen,
        qualifikationen=qualifikationen,
    )


def parse_qualifikationsmatrix(data: bytes, dateiname: str) -> ParsedDatei:
    """Alle Matrix-Blätter einer Bereichsdatei einlesen.

    Quality bringt drei Blätter mit (QM, CS, QS); die anderen Bereiche eines.
    Das Blatt "Qualifikationsdiagramm" enthält nur Formelverweise und wird
    übersprungen.
    """
    wb = load_workbook(io.BytesIO(data), data_only=True)
    ergebnis = ParsedDatei(dateiname=dateiname, matrizen=[])

    for blattname in wb.sheetnames:
        if any(m in blattname.lower() for m in IGNORIERTE_BLAETTER):
            continue
        matrix = _lies_blatt(wb[blattname])
        if matrix is None:
            ergebnis.warnungen.append(
                f"Blatt '{blattname}' übersprungen: keine Matrix erkannt "
                f"(fehlende Kopfzeile 'Anzahl Mitarbeiter' oder keine Personenspalten)."
            )
            continue
        ergebnis.matrizen.append(matrix)

    if not ergebnis.matrizen:
        raise ValueError("In der Datei wurde keine Qualifikationsmatrix gefunden.")
    return ergebnis
