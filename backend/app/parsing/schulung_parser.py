"""Parser für die ``Schulungsübersicht.xlsx``.

Aufbau der Quelldatei (aus den echten Dateien abgeleitet, Stand 06/2025):

* Je Bereich ein Arbeitsblatt: ``betrieblich (gesamt)``, ``Produktion``,
  ``Verwaltung``.
* Die Matrix ist **transponiert**: Spalten sind Mitarbeiter, Zeilen sind
  Schulungen.
* Kopfbereich je Blatt (Spalte E beschriftet, ab Spalte F die Mitarbeiter):
  ``Pers.Nr.`` / ``Name, Vorname`` / ``Abt.``
* Je Schulung **drei aufeinanderfolgende Zeilen**; Spalte B trägt den Turnus,
  Spalte C den Schulungsnamen (nur in der ersten Zeile), Spalte D die Art des
  Werts: ``Initial`` → ``aktuell`` → ``nächste``.

Der Parser ist bewusst tolerant: die Datei ist gewachsen und enthält
Datumsangaben als echtes Datum, als Jahreszahl (``2024``) und als Freitext.
Nicht interpretierbare Werte werden als Warnung gemeldet statt geraten.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO

import openpyxl

#: Turnus-Text → Periode in Monaten. Bewusst konservativ: Spannen
#: ("alle 3 - 5 Jahre") und "bei Bedarf" ergeben KEINE berechenbare Frist.
TURNUS_MONATE: dict[str, int | None] = {
    "jährlich": 12,
    "jaehrlich": 12,
    "alle 2 jahre": 24,
    "alle 2 jahre (und bei bedarf)": 24,
    "alle 3 - 5 jahre": None,
    "alle 3-5 jahre": None,
    "bei bedarf": None,
}

_KOPF_PERSNR = "pers.nr."
_KOPF_NAME = "name, vorname"
_KOPF_ABT = "abt."
_ERSTE_MA_SPALTE = 6  # Spalte F


@dataclass
class ParsedTeilnahme:
    personalnummer: str
    mitarbeiter_name: str | None
    abteilung_kuerzel: str | None
    initial_datum: date | None = None
    aktuell_datum: date | None = None
    naechste_faellig: str | None = None


@dataclass
class ParsedSchulung:
    bereich: str
    name: str
    turnus: str | None
    turnus_monate: int | None
    sort_order: int
    teilnahmen: list[ParsedTeilnahme] = field(default_factory=list)


@dataclass
class ParseResult:
    schulungen: list[ParsedSchulung] = field(default_factory=list)
    warnungen: list[str] = field(default_factory=list)

    @property
    def teilnahmen_gesamt(self) -> int:
        return sum(len(s.teilnahmen) for s in self.schulungen)


def _norm(value: object) -> str:
    """Whitespace (inkl. Zeilenumbrüche in Zellen) auf ein Leerzeichen normieren."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def turnus_zu_monaten(turnus: str | None) -> int | None:
    """Turnus-Freitext in eine Monatsperiode übersetzen (None wenn unbestimmt)."""
    if not turnus:
        return None
    return TURNUS_MONATE.get(_norm(turnus).lower())


def _als_datum(value: object) -> date | None:
    """Zelle als Datum lesen — echtes Datum, Jahreszahl oder ISO-Text."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _norm(value)
    # Reine Jahreszahl ("2024") → 1. Januar, damit die Reihenfolge stimmt.
    if re.fullmatch(r"(19|20)\d{2}", text):
        return date(int(text), 1, 1)
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d.%m.%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _mitarbeiter_spalten(ws) -> tuple[dict[int, tuple[str, str | None, str | None]], list[str]]:
    """Kopfbereich lesen → {Spalte: (Personalnummer, Name, Abteilungskürzel)}."""
    warnungen: list[str] = []
    zeile_nr = zeile_name = zeile_abt = None
    for r in range(1, 12):
        label = _norm(ws.cell(r, 5).value).lower()  # Spalte E
        if label.startswith(_KOPF_PERSNR):
            zeile_nr = r
        elif label.startswith(_KOPF_NAME):
            zeile_name = r
        elif label.startswith(_KOPF_ABT):
            zeile_abt = r

    if zeile_nr is None:
        warnungen.append(f"[{ws.title}] Kopfzeile 'Pers.Nr.' nicht gefunden — Blatt übersprungen.")
        return {}, warnungen

    spalten: dict[int, tuple[str, str | None, str | None]] = {}
    for c in range(_ERSTE_MA_SPALTE, ws.max_column + 1):
        persnr = _norm(ws.cell(zeile_nr, c).value)
        if not persnr:
            continue
        name = _norm(ws.cell(zeile_name, c).value) if zeile_name else ""
        abt = _norm(ws.cell(zeile_abt, c).value) if zeile_abt else ""
        spalten[c] = (persnr, name or None, abt or None)
    return spalten, warnungen


def parse_schulungsuebersicht(data: bytes) -> ParseResult:
    """``Schulungsübersicht.xlsx`` einlesen; wirft nicht, sondern sammelt Warnungen."""
    result = ParseResult()
    wb = openpyxl.load_workbook(BytesIO(data), data_only=True)

    for ws in wb.worksheets:
        bereich = _norm(ws.title).replace(" (gesamt)", "")
        spalten, warn = _mitarbeiter_spalten(ws)
        result.warnungen.extend(warn)
        if not spalten:
            continue

        sort_order = 0
        r = 1
        while r <= ws.max_row:
            name = _norm(ws.cell(r, 3).value)  # Spalte C
            art = _norm(ws.cell(r, 4).value).lower()  # Spalte D
            # Eine Schulung beginnt dort, wo ein Name UND "Initial" stehen.
            if not name or art != "initial":
                r += 1
                continue

            turnus = _norm(ws.cell(r, 2).value) or None  # Spalte B
            if turnus in ("-", "–"):
                turnus = None
            monate = turnus_zu_monaten(turnus)
            if turnus and monate is None and _norm(turnus).lower() not in TURNUS_MONATE:
                result.warnungen.append(
                    f"[{ws.title}] Unbekannter Turnus '{turnus}' bei '{name}' — keine Fälligkeit berechenbar."
                )

            sort_order += 1
            schulung = ParsedSchulung(
                bereich=bereich,
                name=name,
                turnus=turnus,
                turnus_monate=monate,
                sort_order=sort_order,
            )

            # Die beiden Folgezeilen tragen 'aktuell' und 'nächste'.
            zeile_aktuell = r + 1 if _norm(ws.cell(r + 1, 4).value).lower() == "aktuell" else None
            zeile_naechste = r + 2 if _norm(ws.cell(r + 2, 4).value).lower().startswith("näch") else None

            for c, (persnr, ma_name, abt) in spalten.items():
                initial = _als_datum(ws.cell(r, c).value)
                aktuell = _als_datum(ws.cell(zeile_aktuell, c).value) if zeile_aktuell else None
                naechste = _norm(ws.cell(zeile_naechste, c).value) if zeile_naechste else ""
                # Nur Zeilen aufnehmen, die für diesen Mitarbeiter etwas aussagen.
                if initial is None and aktuell is None and not naechste:
                    continue
                schulung.teilnahmen.append(
                    ParsedTeilnahme(
                        personalnummer=persnr,
                        mitarbeiter_name=ma_name,
                        abteilung_kuerzel=abt,
                        initial_datum=initial,
                        aktuell_datum=aktuell,
                        naechste_faellig=naechste or None,
                    )
                )

            result.schulungen.append(schulung)
            r += 3

    return result
