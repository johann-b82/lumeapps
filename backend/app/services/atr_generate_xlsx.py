# backend/app/services/atr_generate_xlsx.py
"""Fill the stored ATR template (frame) with a delivery's matched rows (Phase B).

Template-as-frame: keep the template's header block / table header / totals /
certification formatting; rewrite the part-table region with the delivery items
(grouped by category), copying cell style from a template part row. Unmatched
items get a red fill so the operator fixes them in Excel.
"""
from __future__ import annotations

import asyncio
import re
import shutil
import uuid as _uuid
import zipfile
from copy import copy
from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.worksheet.properties import PageSetupProperties

from app.services.atr_format import normalize_po_pos

_RED = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
_NCOLS = 14  # A..N
_DE_NUM = "[$-407]0.00"  # force German decimal comma regardless of LibreOffice locale
_ACM_ADDRESS = "ACM GmbH - Brandstücken 16 - 22549 Hamburg"
# Certification block (row wraps to 3 lines; merged cells don't auto-size height).
_CERT_ROW_HEIGHT = 48.0


def _de_date(d) -> str:
    """German TT.MM.JJJJ for date/datetime; passthrough for anything else."""
    try:
        return d.strftime("%d.%m.%Y")
    except AttributeError:
        return str(d)


def _visible_sheet(wb):
    vis = [w for w in wb.worksheets if w.sheet_state == "visible"]
    if len(vis) != 1:
        raise ValueError(f"expected one visible sheet, found {len(vis)}")
    return vis[0]


def _find_totals_row(ws, first_part_row: int) -> int:
    for r in range(first_part_row, ws.max_row + 1):
        f = ws.cell(r, 6).value
        if f and "total" in str(f).lower():
            return r
    raise ValueError("ATR template has no 'Total weight' row in column F")


def _find_label_row(ws, col: int, text: str, upto: int = 20) -> int | None:
    """Row where cell(row, col) starts with `text` (case-insensitive). Lets the
    generator locate header fields by label so it works for both the A350 and
    the more compact A380 layout instead of relying on fixed cell addresses."""
    t = text.lower()
    for r in range(1, upto + 1):
        v = ws.cell(r, col).value
        if v and str(v).strip().lower().startswith(t):
            return r
    return None


def _capture_row_styles(ws, row: int) -> list:
    return [copy(ws.cell(row, c)._style) for c in range(1, _NCOLS + 1)]


def build_atr_xlsx(template_bytes: bytes, delivery, items, logo_bytes: bytes | None = None) -> bytes:
    wb = load_workbook(BytesIO(template_bytes))
    ws = _visible_sheet(wb)

    is_a380 = "380" in (delivery.ac_programme or "")
    # Table header row ("PO Pos") located by label so A350 (row 13) and the more
    # compact A380 (row 10) both work; the part region starts one row below.
    header_row = _find_label_row(ws, 1, "PO Pos")
    if header_row is None:
        raise ValueError("ATR template: 'PO Pos' table header row not found")
    first_part_row = header_row + 1

    # --- header block, located by label (template defaults stay otherwise) ---
    sr = _find_label_row(ws, 1, "Supplier:")
    if sr:
        ws.cell(sr, 4, _ACM_ADDRESS)  # supplier address (static)
    br = _find_label_row(ws, 1, "Manufacturing Process Reference")
    if br and delivery.ba_auftrag is not None:
        ws.cell(br, 4, delivery.ba_auftrag)
    pr = _find_label_row(ws, 6, "Purchase Order No")
    if pr and delivery.po_number is not None:
        ws.cell(pr, 7, delivery.po_number)
    mr = _find_label_row(ws, 6, "MSN:")
    if mr:
        if is_a380:
            ws.cell(mr, 7, delivery.msn or "N/A Spare Part")  # A380 spares: no MSN
        elif delivery.msn is not None:
            ws.cell(mr, 7, delivery.msn)
            # A350 PO line middle segment "79 - <MSN> - 94": MSN zero-padded in J1.
            ws.cell(1, 10, str(delivery.msn).zfill(4))
    wr = _find_label_row(ws, 1, "Weighing date")
    if wr:
        if getattr(delivery, "weighing_date", None):
            ws.cell(wr, 3, _de_date(delivery.weighing_date))
        if getattr(delivery, "testing_date", None):
            ws.cell(wr, 12, _de_date(delivery.testing_date))
        if delivery.set_title is not None:  # banner row sits directly above
            ws.cell(wr - 1, 1, delivery.set_title)
    # Print header (right section): Doc-No / Date / Page. The PDF path rebuilds
    # this via UNO (atr_uno_header.py); mirror all three lines here so printing
    # the raw .xlsx from Excel shows the same header. &P/&N are Excel page fields.
    ws.oddHeader.right.text = (
        f"Doc-No.: {atr_doc_no(delivery)}\n"
        f"Date: {date.today().strftime('%d.%m.%Y')}\n"
        "Page: &P of &N"
    )

    # --- capture reference styles BEFORE mutating the region ---
    part_style = _capture_row_styles(ws, first_part_row + 1)  # a part row
    section_style = _capture_row_styles(ws, first_part_row)   # a section header row

    totals_row = _find_totals_row(ws, first_part_row)
    region_count = max(0, totals_row - first_part_row)

    # openpyxl delete_rows/insert_rows move cell CONTENT but leave merged-cell
    # ranges (and their wide legend/certification blocks) where they were.
    # Snapshot the merges now so we can realign them after the row surgery.
    _orig_merges = [(m.min_row, m.min_col, m.max_row, m.max_col)
                    for m in ws.merged_cells.ranges]

    # group items by category in first-seen order
    grouped: list[tuple[str | None, list]] = []
    index: dict[str | None, int] = {}
    for it in items:
        cat = it.category
        if cat not in index:
            index[cat] = len(grouped)
            grouped.append((cat, []))
        grouped[index[cat]][1].append(it)

    # rows we will write: one section header per category + one per item
    out_rows = sum(1 + len(lst) for _, lst in grouped)

    # clear the example region and resize it to out_rows
    if region_count:
        ws.delete_rows(first_part_row, region_count)
    if out_rows:
        ws.insert_rows(first_part_row, out_rows)

    # Realign merged cells to the shifted content: header block (above the part
    # region) stays; merges that were inside the rewritten region are dropped;
    # everything at/below the totals row shifts by the net row delta.
    delta = out_rows - region_count
    ws.merged_cells.ranges = []  # drop all stale ranges; re-add correctly below
    for (r1, c1, r2, c2) in _orig_merges:
        if r1 < first_part_row:
            ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        elif r1 >= totals_row:
            ws.merge_cells(start_row=r1 + delta, start_column=c1,
                           end_row=r2 + delta, end_column=c2)

    r = first_part_row
    total_weight = Decimal("0")
    for cat, lst in grouped:
        # section header row: category label centered across A..H (through
        # the Weight column), then the inspection columns I..N stay separate.
        for c in range(1, _NCOLS + 1):
            ws.cell(r, c)._style = copy(section_style[c - 1])
        ws.cell(r, 1, cat or "")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        ws.cell(r, 1).alignment = Alignment(horizontal="center", vertical="center")
        r += 1
        for it in lst:
            for c in range(1, _NCOLS + 1):
                ws.cell(r, c)._style = copy(part_style[c - 1])
            ws.cell(r, 1, normalize_po_pos(it.po_pos) or "")
            ws.cell(r, 2, it.supplier_article_code or "")
            ws.cell(r, 3, it.part_number or "")
            ws.cell(r, 4, it.part_name or "")
            ws.cell(r, 5, it.serial_numbers or "N/A")
            ws.cell(r, 6, it.drawing_number_issue or "")
            ws.cell(r, 7, it.qty)
            if it.weight_kg is not None:
                # weight_kg is the per-unit weight; the line (and the total) must
                # account for the quantity.
                line_weight = it.weight_kg * (it.qty or 1)
                wc = ws.cell(r, 8, float(line_weight))
                wc.number_format = _DE_NUM
                total_weight += line_weight
            for c, mark in zip(range(9, 14), ("P", "P", "P", "P", "P")):
                ws.cell(r, c, mark)
            ws.cell(r, 14, "OK")
            if it.match_status != "matched":
                for c in range(1, _NCOLS + 1):
                    ws.cell(r, c).fill = _RED
            r += 1

    # totals block shifted down by (out_rows - region_count + region_count) — re-find it
    new_totals_row = _find_totals_row(ws, first_part_row)
    tw = ws.cell(new_totals_row, 8, float(total_weight))
    tw.number_format = _DE_NUM
    # Max. Guaranteed weight on the next totals label row, if present
    for rr in range(new_totals_row, min(new_totals_row + 4, ws.max_row + 1)):
        if "max" in str(ws.cell(rr, 6).value or "").lower() and delivery.max_guaranteed_weight_kg is not None:
            mg = ws.cell(rr, 8, float(delivery.max_guaranteed_weight_kg))
            mg.number_format = _DE_NUM

    # Grow the certification row so its wrapped 3-line text isn't clipped —
    # merged cells (A..F / G..N) don't auto-size row height. Located after the
    # row surgery since the block shifts with the rewritten part region.
    cert_row = next((c.row for row in ws.iter_rows() for c in row
                     if isinstance(c.value, str) and "hereby certify" in c.value), None)
    if cert_row is not None:
        ws.row_dimensions[cert_row].height = _CERT_ROW_HEIGHT

    # QA signer (C/L on the "Date:" row). The template ships a static name;
    # make the delivery's qa_signer authoritative (blank when unset) so the
    # field actually drives the document.
    sig_row = next((r for r in range(new_totals_row, ws.max_row + 1)
                    if str(ws.cell(r, 1).value or "").strip() == "Date:"), None)
    if sig_row is not None:
        ws.cell(sig_row, 3, delivery.qa_signer or "")
        ws.cell(sig_row, 12, delivery.qa_signer or "")
        # The template ships =TODAY() in the two certification "Date:" cells,
        # which LibreOffice renders in its own locale (e.g. 8/5/2026). Overwrite
        # them with a fixed DD.MM.YYYY string so all four date fields on the
        # sheet share one format (matches the weighing/testing dates above).
        cert_date = _de_date(delivery.testing_date or delivery.weighing_date or date.today())
        ws.cell(sig_row, 2, cert_date)   # left  "Date:" (B)
        ws.cell(sig_row, 8, cert_date)   # right "Date:" (H, merged H:I)

    # Fit to one page WIDE (kills the horizontal overflow that doubled the page
    # count); height flows to as many pages as needed. Constrain print_area to
    # A..N so stray far-right cells don't spawn extra pages.
    ws.print_area = f"A1:N{ws.max_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    # Symmetric L/R margins so the right-aligned print header (Doc-No / Date /
    # Page) lines up with the right edge of the fit-to-width table box below.
    # The template ships asymmetric margins (left > right), which pushes the
    # header past the box. Horizontal fit depends only on the fixed A..N columns,
    # so this alignment holds for every delivery regardless of page count.
    ws.page_margins.right = ws.page_margins.left
    ws.page_setup.scale = None  # scale and fitToPage are mutually exclusive
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

    bio = BytesIO()
    wb.save(bio)
    return _inject_header_logo(bio.getvalue(), template_bytes, logo_bytes)


# --- print-header logo restore -------------------------------------------------
# openpyxl keeps the header '&G' code but drops the backing image parts on save,
# so a printed .xlsx shows no logo. Copy those parts back from the template
# (which is wired correctly) and re-add the <legacyDrawingHF> sheet reference.
_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


def _inject_header_logo(xlsx_bytes: bytes, template_bytes: bytes,
                        logo_bytes: bytes | None = None) -> bytes:
    try:
        ztpl = zipfile.ZipFile(BytesIO(template_bytes))
        media = ztpl.read("xl/media/image1.png")
        vml = ztpl.read("xl/drawings/vmlDrawing1.vml")
        vml_rels = ztpl.read("xl/drawings/_rels/vmlDrawing1.vml.rels")
    except KeyError:
        return xlsx_bytes  # template has no header logo → nothing to restore
    # Use the configured logo (same mark as the PDF) instead of the template's
    # own, so Excel and PDF show the identical logo. The template's VML shape
    # sizes it (id="LH", ~91x32pt); we only swap the image bytes it points to.
    if logo_bytes:
        swapped = _excel_logo_png(logo_bytes)
        if swapped:
            media = swapped

    zin = zipfile.ZipFile(BytesIO(xlsx_bytes))
    names = zin.namelist()
    target = next((n for n in names
                   if n.startswith("xl/worksheets/sheet") and n.endswith(".xml")
                   and b"&amp;G" in zin.read(n)), None)
    if target is None:
        return xlsx_bytes  # no header-graphic sheet found
    rels_name = f"xl/worksheets/_rels/{target.split('/')[-1]}.rels"
    have_rels = rels_name in names

    out = BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        for n in names:
            data = zin.read(n)
            if n == "[Content_Types].xml":
                data = _ct_with_defaults(data)
            elif n == target:
                data = _sheet_with_legacy_hf(data)
            elif n == rels_name:
                data = _rels_with_vml(data)
            zout.writestr(n, data)
        zout.writestr("xl/media/image1.png", media)
        zout.writestr("xl/drawings/vmlDrawing1.vml", vml)
        zout.writestr("xl/drawings/_rels/vmlDrawing1.vml.rels", vml_rels)
        if not have_rels:
            zout.writestr(rels_name,
                          '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                          f'<Relationships xmlns="{_REL_NS}">'
                          f'<Relationship Id="rIdHF1" Type="{_R_NS}/vmlDrawing" '
                          'Target="../drawings/vmlDrawing1.vml"/></Relationships>')
    return out.getvalue()


def _excel_logo_png(logo_bytes: bytes) -> bytes | None:
    """Crop transparent borders off the configured logo and cap its width for
    the Excel header image (the VML shape controls its on-page size). Returns
    None on any failure so the template's own logo is kept as a fallback."""
    try:
        from PIL import Image
        im = Image.open(BytesIO(logo_bytes)).convert("RGBA")
        bbox = im.getbbox()
        if bbox:
            im = im.crop(bbox)
        if im.width > 800:
            im = im.resize((800, max(1, round(im.height * 800 / im.width))))
        out = BytesIO()
        im.save(out, format="PNG")
        return out.getvalue()
    except Exception:  # noqa: BLE001
        return None


def _ct_with_defaults(ct: bytes) -> bytes:
    s = ct.decode("utf-8")
    add = ""
    if 'Extension="png"' not in s:
        add += '<Default Extension="png" ContentType="image/png"/>'
    if 'Extension="vml"' not in s:
        add += ('<Default Extension="vml" '
                'ContentType="application/vnd.openxmlformats-officedocument.vmlDrawing"/>')
    return (re.sub(r"(<Types[^>]*>)", r"\1" + add, s, count=1) if add else s).encode("utf-8")


def _sheet_with_legacy_hf(sheet: bytes) -> bytes:
    s = sheet.decode("utf-8")
    if "legacyDrawingHF" in s:
        return sheet
    if "xmlns:r=" not in s.split(">", 1)[0]:
        s = s.replace("<worksheet ", f'<worksheet xmlns:r="{_R_NS}" ', 1)
    return s.replace("</worksheet>",
                     '<legacyDrawingHF r:id="rIdHF1"/></worksheet>', 1).encode("utf-8")


def _rels_with_vml(rels: bytes) -> bytes:
    s = rels.decode("utf-8")
    if "vmlDrawing" in s:
        return rels
    inject = (f'<Relationship Id="rIdHF1" Type="{_R_NS}/vmlDrawing" '
              'Target="../drawings/vmlDrawing1.vml"/>')
    return s.replace("</Relationships>", inject + "</Relationships>", 1).encode("utf-8")


# Serialize LibreOffice across the single-worker api container (mirror signage_pptx).
_LO_SEMAPHORE = asyncio.Semaphore(1)
_LO_TIMEOUT_S = 60


# Fallback Doc-No for this report family; the header date is the report
# generation date. Applied via LibreOffice UNO because openpyxl mangles the
# print header — see atr_uno_header.py.
_ATR_DOC_NO = "ACM-A350CRC-ATR-4545-01 / Issue: 01"
_UNO_SCRIPT = str(Path(__file__).with_name("atr_uno_header.py"))


def atr_doc_no(delivery) -> str:
    """Print-header Doc-No, derived from the delivery's ATR number
    (``4820`` -> ``ACM-A350CRC-ATR-4820-01 / Issue: 01``). Falls back to the
    family constant when the ATR number is unset."""
    n = (getattr(delivery, "atr_number", None) or "").strip()
    if not n:
        return _ATR_DOC_NO
    if "380" in (getattr(delivery, "ac_programme", None) or ""):
        return f"ACM-A380-ATR-{n}-01 / Issue: 01"
    return f"ACM-A350CRC-ATR-{n}-01 / Issue: 01"


_LOGO_HEIGHT_MM = 8.5  # fits the header band fully (incl. the "AEROSPACE"
#                        tagline); taller clips at the band bottom.
_LOGO_DPI = 300        # LibreOffice honours the PNG DPI for the header
#                        background graphic, so render at print resolution.


def _prep_header_logo(logo_bytes: bytes) -> bytes:
    """Crop transparent borders and scale the logo to a fixed physical height at
    print DPI, so LibreOffice's header background graphic (placed at native size
    = pixels / DPI, top-left) shows the whole mark, sharp and unclipped. Falls
    back to the raw bytes if Pillow is unavailable or the image can't be read."""
    try:
        from PIL import Image
        im = Image.open(BytesIO(logo_bytes)).convert("RGBA")
        bbox = im.getbbox()
        if bbox:
            im = im.crop(bbox)
        h = max(1, round(_LOGO_HEIGHT_MM / 25.4 * _LOGO_DPI))
        w = max(1, round(im.width * h / im.height))
        out = BytesIO()
        im.resize((w, h)).save(out, format="PNG", dpi=(_LOGO_DPI, _LOGO_DPI))
        return out.getvalue()
    except Exception:  # noqa: BLE001 — never fail generation over the logo
        return logo_bytes


async def convert_xlsx_to_pdf(xlsx_bytes: bytes, doc_no: str = _ATR_DOC_NO,
                              date_str: str | None = None,
                              logo_bytes: bytes | None = None,
                              logo_ext: str = "png") -> bytes:
    if date_str is None:
        date_str = date.today().strftime("%d.%m.%Y")
    async with _LO_SEMAPHORE:
        tempdir = Path(f"/tmp/atr_{_uuid.uuid4()}")
        try:
            tempdir.mkdir(parents=True, exist_ok=True)
            src = tempdir / "atr.xlsx"
            src.write_bytes(xlsx_bytes)
            out = tempdir / "atr.pdf"
            # openpyxl drops the template's header graphic (keeps the '&G'
            # placeholder), so re-supply the logo to the UNO step, which anchors
            # it as a floating shape in the header band (atr_uno_header.py).
            args = ["/usr/bin/python3", _UNO_SCRIPT, str(src), str(out), doc_no, date_str]
            if logo_bytes:
                if logo_ext != "svg":
                    logo_bytes = _prep_header_logo(logo_bytes)
                    logo_ext = "png"
                logo = tempdir / f"logo.{logo_ext}"
                logo.write_bytes(logo_bytes)
                args.append(str(logo))
            proc = await asyncio.create_subprocess_exec(
                *args,
                stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE,
            )
            try:
                _, err = await asyncio.wait_for(proc.communicate(), timeout=_LO_TIMEOUT_S)
            except asyncio.TimeoutError as exc:
                try:
                    proc.kill()
                except ProcessLookupError:
                    pass
                raise RuntimeError("xlsx->pdf conversion timed out") from exc
            if proc.returncode != 0 or not out.exists():
                raise RuntimeError(
                    f"uno header/pdf failed: {err.decode('utf-8', 'replace')[-500:]}"
                )
            return out.read_bytes()
        finally:
            shutil.rmtree(tempdir, ignore_errors=True)
