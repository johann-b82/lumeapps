"""Schulungsübersicht je Person als PDF — Formblatt 71 (v1.91).

Bildet das bestehende Papierformular nach: Kopf mit Name und Funktion, darunter
je Schulung eine Zeile mit laufender Nummer, Zeitraum, Bezeichnung, IN/EX-Kreuz
und dem Feld "Schulungsnachweis vorhanden".

Aufbau wie beim Wartungsnachweis (``maintenance_pdf``): openpyxl schreibt eine
.xlsx, LibreOffice headless macht daraus ein PDF. Kein UNO nötig — alles steht
in Zellen.

Für einen neuen Mitarbeiter sind Zeitraum und Nachweis-Spalte leer: das Blatt
ist dann der Plan, den er abarbeitet, und wird beim Absolvieren handschriftlich
oder über die Anwendung nachgeführt.
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.page import PageMargins

from app.services.maintenance_pdf import convert_xlsx_to_pdf
from app.services.pdf_logo import LogoBild, bild_einsetzen

#: Kopfangaben des Formblatts. Stehen so auf dem Papierformular und ändern sich
#: nur mit einer neuen Revision des Formblatts selbst.
FORMBLATT = "Formblatt 71"
REVISIONS_INDEX = "A"
REVISIONS_STAND = "22.03.2022"
AUSGABEDATUM = "22.03.2022"

#: Spaltennummern der vier Ankreuzfelder (D-G).
SP_IN, SP_EX, SP_JA, SP_NEIN = 4, 5, 6, 7

_THIN = Side(style="thin", color="000000")
_RAHMEN = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_LINKS_OBEN = Alignment(horizontal="left", vertical="top", wrap_text=True)
_MITTE = Alignment(horizontal="center", vertical="center")
_MITTE_OBEN = Alignment(horizontal="center", vertical="top")


@dataclass
class UebersichtZeile:
    """Eine Zeile des Formblatts."""

    bezeichnung: str
    #: Freitext wie "16.03.2026 - 17.03.2026"; leer, solange nicht absolviert.
    zeitraum: str = ""
    #: Anbieter samt Anschrift, erscheint unter der Bezeichnung.
    anbieter: str = ""
    #: True = intern (IN), False = extern (EX), None = noch offen.
    intern: bool | None = None
    #: True/False setzt das Kreuz in "Schulungsnachweis vorhanden".
    nachweis: bool | None = None


def _kopf(ws, name: str, funktion: str, logo: LogoBild | None = None) -> int:
    """Titelblock und Personenangaben. Gibt die nächste freie Zeile zurück.

    Mit Logo entsteht oben ein Logo-Band; Titel und Formblatt-Block rücken
    darunter. Ohne Logo bleibt das Layout wie zuvor (Titel ab Zeile 1).
    """
    # "Blatt x von y" steht als Seitenkopf (&P/&N) — fest verdrahtet wäre es
    # falsch, sobald die Liste auf eine zweite Seite läuft.
    ws.oddHeader.right.text = "&9Blatt &P von &N"
    ws.evenHeader.right.text = ws.oddHeader.right.text

    top = 1
    if logo is not None:
        bild_einsetzen(ws, logo, "A1")
        for r in (1, 2, 3):
            ws.row_dimensions[r].height = 20
        top = 4  # Titel/Formblatt beginnen unter dem Logo-Band

    # Formblatt-Block rechts (C:G zusammengefasst), auf Höhe des Titels.
    for i, text in enumerate(
        [
            FORMBLATT,
            f"Revisions-Index: {REVISIONS_INDEX}",
            f"Revisions-Stand: {REVISIONS_STAND}",
        ]
    ):
        zeile = top + i
        ws.merge_cells(start_row=zeile, start_column=3, end_row=zeile, end_column=SP_NEIN)
        zelle = ws.cell(row=zeile, column=3, value=text)
        zelle.font = Font(size=9)
        zelle.alignment = Alignment(horizontal="right", vertical="center")

    # Titel links, über A:B.
    titel = ws.cell(row=top, column=1, value="Schulungsübersicht")
    titel.font = Font(size=16, bold=True)
    titel.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=top, start_column=1, end_row=top + 1, end_column=2)

    nz = top + 4
    ws.cell(row=nz, column=1, value="Name:").font = Font(bold=True)
    ws.cell(row=nz, column=2, value=name)
    ws.cell(row=nz + 1, column=1, value="Funktion:").font = Font(bold=True)
    ws.cell(row=nz + 1, column=2, value=funktion or "—")
    for zeile in (nz, nz + 1):
        ws.merge_cells(start_row=zeile, start_column=2, end_row=zeile, end_column=SP_NEIN)
    return nz + 2


def _tabellenkopf(ws, zeile: int) -> int:
    """Zweizeiliger Tabellenkopf wie auf dem Formular.

    IN/EX und ja/nein sind vier eigene Ankreuzfelder, nicht zwei kombinierte —
    das Blatt wird ausgedruckt und von Hand abgehakt.
    """
    ws.cell(row=zeile, column=1, value="Laufende\nNummer:")
    ws.cell(row=zeile, column=2, value="durchgeführt am/\nvon … bis:")
    ws.cell(row=zeile, column=3, value="Bezeichnung der Aus- und Fortbildungsmaßnahme")
    ws.cell(row=zeile, column=SP_IN, value="Schulungsnachweis\nvorhanden")
    ws.merge_cells(start_row=zeile, start_column=SP_IN, end_row=zeile, end_column=SP_NEIN)

    unten = zeile + 1
    ws.cell(row=unten, column=3, value="Interne (IN) oder externe (EX) Maßnahme:")
    ws.cell(row=unten, column=SP_IN, value="IN")
    ws.cell(row=unten, column=SP_EX, value="EX")
    ws.cell(row=unten, column=SP_JA, value="ja")
    ws.cell(row=unten, column=SP_NEIN, value="nein")
    for spalte in (1, 2):
        ws.merge_cells(
            start_row=zeile, start_column=spalte, end_row=unten, end_column=spalte
        )

    for r in (zeile, unten):
        for c in range(1, SP_NEIN + 1):
            zelle = ws.cell(row=r, column=c)
            zelle.font = Font(size=9, bold=True)
            # wrap_text auch in den zentrierten Spalten, sonst läuft
            # "Schulungsnachweis vorhanden" in einer Zeile durch.
            zelle.alignment = (
                Alignment(horizontal="center", vertical="center", wrap_text=True)
                if c >= SP_IN
                else _LINKS_OBEN
            )
            zelle.border = _RAHMEN
    ws.row_dimensions[zeile].height = 32
    return unten + 1


def _zeile_schreiben(ws, r: int, nr: int, z: UebersichtZeile) -> None:
    ws.cell(row=r, column=1, value=f"{nr:02d}")
    ws.cell(row=r, column=2, value=z.zeitraum)
    # Anbieter kommt als zweiter Absatz in dieselbe Zelle — auf dem Formular
    # steht er unter der Bezeichnung, nicht in einer eigenen Spalte.
    text = z.bezeichnung + (f"\n\n{z.anbieter}" if z.anbieter else "")
    ws.cell(row=r, column=3, value=text)
    # Kreuz nur dort, wo die Angabe feststeht; offen bleibt leer zum Ankreuzen.
    ws.cell(row=r, column=SP_IN, value="X" if z.intern is True else "")
    ws.cell(row=r, column=SP_EX, value="X" if z.intern is False else "")
    ws.cell(row=r, column=SP_JA, value="X" if z.nachweis is True else "")
    ws.cell(row=r, column=SP_NEIN, value="X" if z.nachweis is False else "")

    for c in range(1, SP_NEIN + 1):
        zelle = ws.cell(row=r, column=c)
        zelle.border = _RAHMEN
        zelle.font = Font(size=9)
        # Nummer und Zeitraum stehen oben wie im Original, die Kreuzfelder
        # mittig — dort wird von Hand ein X gesetzt.
        zelle.alignment = _MITTE_OBEN if c == 1 else (
            _MITTE if c >= SP_IN else _LINKS_OBEN
        )
    # Zwei Textzeilen Platz; mit Anbieteranschrift entsprechend mehr. Enger als
    # zuvor, damit auf ein Blatt mehr Schulungen passen.
    ws.row_dimensions[r].height = 38 if z.anbieter else 24


def _fuss(ws, freigegeben_von: str, erstellt_von: str) -> None:
    """Fußblock als echte Seitenfußzeile.

    Nicht als Tabellenzeilen unter der letzten Schulung: dort klebte er am
    Tabellenende statt am Blattfuß, und bei wenigen Zeilen stand er mitten auf
    der Seite. Als Fußzeile sitzt er immer unten — und wiederholt sich, wenn
    die Liste auf eine zweite Seite läuft.

    ``&9`` setzt die Schriftgröße.

    Bewusst je Abschnitt EINE Zeile: openpyxl kodiert ein "\\n" in der Fußzeile
    als OOXML-Escape ``_x000a_``, und LibreOffice gibt das wörtlich aus statt
    umzubrechen. Die vier Felder des Formblatts stehen daher nebeneinander
    statt zweizeilig untereinander.
    """
    ws.oddFooter.left.text = (
        f"&9Aktualisiert am: ________     Freigegeben von: {freigegeben_von or '________'}"
    )
    ws.oddFooter.right.text = (
        f"&9Ausgabedatum: {AUSGABEDATUM}     Erstellt von: {erstellt_von or '________'}"
    )
    # Gleiche Fußzeile auf Folgeseiten.
    ws.evenFooter.left.text = ws.oddFooter.left.text
    ws.evenFooter.right.text = ws.oddFooter.right.text


def fuelle_blatt(
    ws,
    name: str,
    funktion: str,
    zeilen: list[UebersichtZeile],
    freigegeben_von: str = "",
    erstellt_von: str = "",
    logo: LogoBild | None = None,
) -> None:
    """Formblatt 71 in ein vorhandenes Arbeitsblatt schreiben.

    Herausgezogen aus ``baue_xlsx``, damit das Blatt auch als zweite Seite eines
    kombinierten Onboarding-Pakets (mit dem Einarbeitungsplan) dienen kann.
    """
    ws.title = "Schulungsübersicht"
    for spalte, breite in zip("ABCDEFG", (11, 16, 58, 5, 5, 5, 6)):
        ws.column_dimensions[spalte].width = breite

    r = _kopf(ws, name, funktion, logo)
    kopfzeile = r
    r = _tabellenkopf(ws, r)
    for i, z in enumerate(zeilen, start=1):
        _zeile_schreiben(ws, r, i, z)
        r += 1
    _fuss(ws, freigegeben_von, erstellt_von)

    ws.print_area = f"A1:G{r - 1}"
    ws.page_setup.orientation = "portrait"
    # A4, nicht der LibreOffice-Standard Letter — das Formblatt wird in
    # Deutschland gedruckt und abgeheftet.
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    # Unten mehr Rand, damit die Fußzeile nicht an der Tabelle klebt.
    ws.page_margins = PageMargins(
        left=0.5, right=0.4, top=0.6, bottom=0.8, header=0.3, footer=0.35
    )
    # Läuft die Liste auf eine zweite Seite, wiederholt sich der Tabellenkopf.
    ws.print_title_rows = f"{kopfzeile}:{kopfzeile + 1}"


def baue_xlsx(
    name: str,
    funktion: str,
    zeilen: list[UebersichtZeile],
    freigegeben_von: str,
    erstellt_von: str,
    logo: LogoBild | None = None,
) -> bytes:
    """Formblatt 71 als .xlsx (Zwischenschritt zur PDF-Erzeugung)."""
    wb = Workbook()
    fuelle_blatt(wb.active, name, funktion, zeilen, freigegeben_von, erstellt_von, logo)
    puffer = BytesIO()
    wb.save(puffer)
    return puffer.getvalue()


async def erzeuge_schulungsuebersicht_pdf(
    name: str,
    funktion: str,
    zeilen: list[UebersichtZeile],
    freigegeben_von: str = "",
    erstellt_von: str = "",
    logo: LogoBild | None = None,
) -> bytes:
    return await convert_xlsx_to_pdf(
        baue_xlsx(name, funktion, zeilen, freigegeben_von, erstellt_von, logo)
    )


def dateiname(name: str, stand: date) -> str:
    """Namensschema der Vorlage: 2026.03.03_Marcel_Brose_Schulungsübersicht."""
    teil = "_".join(name.split()) or "Unbekannt"
    return f"{stand:%Y.%m.%d}_{teil}_Schulungsuebersicht"


# ---------------------------------------------------------------------------
# Vorgangs-Variante (v1.108): dasselbe Formblatt 71 mit QR-Code + Feld-Layout,
# damit ein ausgefüllter Scan über den QR zugeordnet und halbautomatisch geprüft
# werden kann (Prüfung teilt sich app.services.einarbeitung_pruefung). Ohne QR
# bleibt die klassische Schulungsübersicht (oben) unverändert.
# ---------------------------------------------------------------------------

#: Schmalere Spalten als die klassische Übersicht, damit A1:G bei 100 % (ohne
#: Druckskalierung) auf A4 passt → exakte Zeilenhöhen für das Feld-Layout.
_VG_SPALTEN = (11, 15, 42, 5, 5, 5, 6)  # A-G, Summe 89 (passt auf A4-Breite)
_A4_W_PT, _A4_H_PT = 595.276, 841.890
_RAND_L_PX, _RAND_T_PT = 48.0, 36.0  # 0,5" links / oben
_PX_PRO_BREITE, _PX_PAD, _PT_PRO_PX = 7.0, 5.0, 0.75
_VG_QR_PX = 46
_VG_QR_ANKER, _VG_QR_COL, _VG_QR_ROW = "F1", 6, 1


def _vg_x_norm(col: int) -> float:
    px = _RAND_L_PX
    for k in range(1, col):
        px += _VG_SPALTEN[k - 1] * _PX_PRO_BREITE + _PX_PAD
    return (px * _PT_PRO_PX) / _A4_W_PT


def _vg_y_norm(ws, row: int) -> float:
    pt = _RAND_T_PT
    for i in range(1, row):
        h = ws.row_dimensions[i].height
        pt += h if h is not None else 15.0
    return pt / _A4_H_PT


def _vg_norm_box(ws, row: int, c1: int, c2: int, rows: int = 1) -> list[float]:
    return [
        round(_vg_x_norm(c1), 5),
        round(_vg_y_norm(ws, row), 5),
        round(_vg_x_norm(c2 + 1), 5),
        round(_vg_y_norm(ws, row + rows), 5),
    ]


def _vg_qr_png(doc_uid: str) -> bytes:
    import qrcode

    qr = qrcode.QRCode(
        error_correction=qrcode.constants.ERROR_CORRECT_Q, box_size=8, border=2
    )
    qr.add_data(doc_uid)
    qr.make(fit=True)
    puffer = BytesIO()
    qr.make_image(fill_color="black", back_color="white").convert("RGB").save(puffer, "PNG")
    return puffer.getvalue()


def _vg_qr_einsetzen(ws, doc_uid: str) -> None:
    from openpyxl.drawing.image import Image as XLImage

    img = XLImage(BytesIO(_vg_qr_png(doc_uid)))
    img.width = img.height = _VG_QR_PX
    ws.add_image(img, _VG_QR_ANKER)


def _vg_qr_box(ws) -> list[float]:
    x0, y0 = _vg_x_norm(_VG_QR_COL), _vg_y_norm(ws, _VG_QR_ROW)
    return [
        round(x0, 5),
        round(y0, 5),
        round(x0 + (_VG_QR_PX * _PT_PRO_PX) / _A4_W_PT, 5),
        round(y0 + (_VG_QR_PX * _PT_PRO_PX) / _A4_H_PT, 5),
    ]


_VG_MARKE_PX = 16


def _vg_marke_einsetzen(ws, anker: str) -> None:
    from PIL import Image as PILImage
    from openpyxl.drawing.image import Image as XLImage

    puffer = BytesIO()
    PILImage.new("RGB", (40, 40), (0, 0, 0)).save(puffer, "PNG")
    img = XLImage(puffer)
    img.width = img.height = _VG_MARKE_PX
    ws.add_image(img, anker)


def _vg_marke_box(ws, col: int, row: int) -> list[float]:
    x0, y0 = _vg_x_norm(col), _vg_y_norm(ws, row)
    return [
        round(x0, 5),
        round(y0, 5),
        round(x0 + (_VG_MARKE_PX * _PT_PRO_PX) / _A4_W_PT, 5),
        round(y0 + (_VG_MARKE_PX * _PT_PRO_PX) / _A4_H_PT, 5),
    ]


def fuelle_vorgang_blatt(
    ws,
    name: str,
    funktion: str,
    zeilen: list[UebersichtZeile],
    doc_uid: str,
    freigegeben_von: str = "",
    erstellt_von: str = "",
    logo: LogoBild | None = None,
) -> dict:
    """Formblatt 71 als Vorgang (mit QR + Feld-Layout). Gibt das Layout zurück."""
    ws.title = "Schulungsübersicht"
    for spalte, breite in zip("ABCDEFG", _VG_SPALTEN):
        ws.column_dimensions[spalte].width = breite

    r = _kopf(ws, name, funktion, logo)
    kopfzeile = r
    r = _tabellenkopf(ws, r)

    felder: list[dict] = []
    for i, z in enumerate(zeilen, start=1):
        _zeile_schreiben(ws, r, i, z)
        kurz = (z.bezeichnung or "").strip()[:32]
        # Pflichtfelder je Schulungszeile: „durchgeführt am" (Spalte B) und das
        # Nachweis-Kreuz ja/nein (Spalten F–G).
        felder.append({"key": f"durchgefuehrt_{i - 1}", "label": f"Durchgeführt am: {kurz}",
                       "box": _vg_norm_box(ws, r, 2, 2)})
        felder.append({"key": f"nachweis_{i - 1}", "label": f"Nachweis vorhanden: {kurz}",
                       "box": _vg_norm_box(ws, r, SP_JA, SP_NEIN)})
        r += 1
    _fuss(ws, freigegeben_von, erstellt_von)

    _vg_qr_einsetzen(ws, doc_uid)

    # Zwei Passermarken im leeren Bereich unter der Tabelle (unten links/rechts);
    # zusammen mit dem QR oben rechts drei über die Seite verteilte Referenzpunkte.
    marken_zeile = r + 3
    for rr in range(r, marken_zeile + 1):
        if ws.row_dimensions[rr].height is None:
            ws.row_dimensions[rr].height = 15
    _vg_marke_einsetzen(ws, f"A{marken_zeile}")
    _vg_marke_einsetzen(ws, f"G{marken_zeile}")
    marken = [_vg_marke_box(ws, 1, marken_zeile), _vg_marke_box(ws, 7, marken_zeile)]

    ws.print_area = f"A1:G{marken_zeile}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.scale = 100  # keine Skalierung → exakte Zeilenhöhen
    from openpyxl.worksheet.properties import PageSetupProperties

    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=False)
    ws.page_margins = PageMargins(
        left=0.5, right=0.4, top=0.5, bottom=0.6, header=0.3, footer=0.3
    )
    ws.print_title_rows = f"{kopfzeile}:{kopfzeile + 1}"

    return {
        "seite": {"w_pt": _A4_W_PT, "h_pt": _A4_H_PT},
        "qr": {"doc_uid": doc_uid, "box": _vg_qr_box(ws)},
        "marken": marken,
        "felder": felder,
    }


async def erzeuge_schulung_vorgang_pdf(
    name: str,
    funktion: str,
    zeilen: list[UebersichtZeile],
    doc_uid: str,
    freigegeben_von: str = "",
    erstellt_von: str = "",
    logo: LogoBild | None = None,
) -> tuple[bytes, dict]:
    """PDF mit QR + seitenrelativem Feld-Layout für den Schulungsvorgang."""
    wb = Workbook()
    layout = fuelle_vorgang_blatt(
        wb.active, name, funktion, zeilen, doc_uid, freigegeben_von, erstellt_von, logo
    )
    puffer = BytesIO()
    wb.save(puffer)
    return await convert_xlsx_to_pdf(puffer.getvalue()), layout
