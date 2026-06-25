"""Parse a Diehl ATR reference workbook into header defaults + parts.

Pure parsing, no DB. See the design spec's "Source-file analysis" for the
fixed cell map. Reads only the single VISIBLE sheet.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from io import BytesIO

from openpyxl import load_workbook


@dataclass
class ParsedPart:
    row_order: int
    category: str | None
    supplier_article_code: str | None
    part_number: str
    part_number_norm: str
    part_name: str | None
    serial_number: str | None
    drawing_number_issue: str | None
    qty: int
    default_weight_kg: Decimal | None


@dataclass
class ParsedHeader:
    customer: str | None = None
    ac_programme: str | None = None
    work_package: str | None = None
    purchaser_spec: str | None = None
    atp: str | None = None
    supplier_spec: str | None = None
    reference_no: str | None = None
    supplier: str | None = None
    customer_spec: str | None = None
    nscm_code: str | None = None
    ata_chapter: str | None = None
    weighing_equipment: str | None = None


@dataclass
class ParsedWorkbook:
    source_filename: str
    header: ParsedHeader
    parts: list[ParsedPart] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def norm_partno(s: str) -> str:
    return "".join(ch for ch in str(s) if ch.isdigit())


def _cell(ws, coord: str) -> str | None:
    v = ws[coord].value
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _to_decimal(v) -> Decimal | None:
    if v is None:
        return None
    s = str(v).strip().replace(",", ".")
    if not s:
        return None
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _to_int(v, default: int = 1) -> int:
    if v is None:
        return default
    try:
        return int(float(str(v).strip().replace(",", ".")))
    except (ValueError, TypeError):
        return default


def parse_workbook(file_bytes: bytes, source_filename: str) -> ParsedWorkbook:
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    visible = [ws for ws in wb.worksheets if ws.sheet_state == "visible"]
    if len(visible) != 1:
        raise ValueError(
            f"expected exactly one visible sheet, found {len(visible)}"
        )
    ws = visible[0]

    # Layout sanity: row-13 header must look like the ATR table.
    c13 = (_cell(ws, "C13") or "").lower()
    h13 = (_cell(ws, "H13") or "").lower()
    if "part number" not in c13 or "weight" not in h13:
        raise ValueError("unrecognized ATR layout (row-13 header mismatch)")

    header = ParsedHeader(
        customer=_cell(ws, "D1"),
        ac_programme=_cell(ws, "D2"),
        work_package=_cell(ws, "D3"),
        purchaser_spec=_cell(ws, "D4"),
        atp=_cell(ws, "D5"),
        supplier_spec=_cell(ws, "D6"),
        reference_no=_cell(ws, "D7"),
        supplier=_cell(ws, "D8"),
        customer_spec=_cell(ws, "G8"),
        nscm_code=_cell(ws, "G4"),
        ata_chapter=_cell(ws, "G3"),
        weighing_equipment=_cell(ws, "F12"),
    )

    parts: list[ParsedPart] = []
    warnings: list[str] = []
    category: str | None = None
    order = 0
    for r in range(14, ws.max_row + 1):
        a = _cell(ws, f"A{r}")
        c = _cell(ws, f"C{r}")
        f = _cell(ws, f"F{r}")
        if f and "total" in f.lower():
            break  # totals block — stop scanning parts
        if c and c.upper().startswith("VR"):
            weight = _to_decimal(ws[f"H{r}"].value)
            if ws[f"H{r}"].value not in (None, "") and weight is None:
                warnings.append(f"row {r}: unparseable weight {ws[f'H{r}'].value!r}")
            order += 1
            parts.append(ParsedPart(
                row_order=order,
                category=category,
                supplier_article_code=_cell(ws, f"B{r}"),
                part_number=c,
                part_number_norm=norm_partno(c),
                part_name=_cell(ws, f"D{r}"),
                serial_number=_cell(ws, f"E{r}"),
                drawing_number_issue=f,
                qty=_to_int(ws[f"G{r}"].value),
                default_weight_kg=weight,
            ))
        elif a and not c:
            if a.lower().startswith("if applicable"):
                continue
            category = a  # section header

    if not parts:
        warnings.append("no part rows found")
    return ParsedWorkbook(source_filename=source_filename, header=header,
                          parts=parts, warnings=warnings)
