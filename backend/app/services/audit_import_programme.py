"""Import the internal audit programme spreadsheet into the Audit-Modul.

Source: "Internes Auditprogramm 2025-2027" — a matrix where each column is one
audit and each row an attribute (title, abstract, divisions, norm clauses,
target date, auditor). ``parse_programme`` turns that into plain dicts;
``import_programme`` writes them.

Two properties matter more than speed here:

*Idempotent.* Re-running must not duplicate anything. Audits are matched on
``audit_number`` and norm references on (regulation, revision, clause); existing
rows are left untouched and reported as skipped. A half-finished run can simply
be repeated.

*Never auto-advancing.* Everything is imported as ``geplant``, including the six
audits the sheet records as already executed. The module forbids closing an
audit without a human decision, and a bulk importer is not that decision — the
actual execution date and auditor are written into the "Audit durchgeführt"
phase comment so the information is present when a human sets the real status.

Every write is recorded in the audit trail with the importing user as actor.
"""
from __future__ import annotations

import datetime
import re
from dataclasses import dataclass, field
from typing import Any

import sqlalchemy as sa
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import (
    Audit,
    AuditCategoryLink,
    AuditNormLink,
    AuditNormReference,
    AuditPhase,
    AuditPhaseTemplate,
    AuditPhaseTemplateStep,
)
from app.schemas import CurrentUser
from app.services.audit_trail import record

# Matrix layout of the source workbook.
SHEET = "2025-2027"
YEAR_COLUMNS = [("Jahr 1", "BCDEFGHI"), ("Jahr 2", "JKLMNO"), ("Jahr 3", "PQRSTUV")]
ROW = {
    "nr": 3, "title": 4, "abstract": 5, "process": 6, "product": 7,
    "target": 29, "auditor": 30, "actual_date": 32, "actual_auditor": 33,
}
DIVISION_ROWS = range(9, 21)
REFERENCE_ROWS = {
    22: "QMH",
    23: "EASA Part 21G",
    24: "EASA Part 145",
    25: "EN 9100",
    26: "EN 9110",
}
# Revision recorded for each seeded regulation, mirroring the Normmatrix seed.
REFERENCE_REVISION = {
    "QMH": "Rev. 8A",
    "EASA Part 21G": "",
    "EASA Part 145": "",
    "EN 9100": "2018",
    "EN 9110": "2018",
}
# The clause prefix the sheet omits because it is implied by the row label.
CLAUSE_PREFIX = {"EASA Part 21G": "21.A.", "EASA Part 145": "145.A.", "QMH": "QMH "}

# The phase whose comment carries the recorded execution details.
EXECUTION_PHASE_TITLE = "Audit durchgeführt"


@dataclass
class ParsedAudit:
    audit_number: str
    title: str
    objective: str
    categories: list[str]
    scope_label: str
    lead_auditor: str | None
    planned_start: datetime.date | None
    actual_date_raw: str
    actual_auditor: str
    norm_clauses: list[tuple[str, str, str]] = field(default_factory=list)


@dataclass
class ImportReport:
    audits_created: list[str] = field(default_factory=list)
    audits_skipped: list[str] = field(default_factory=list)
    norms_created: list[str] = field(default_factory=list)
    norms_reused: int = 0
    warnings: list[str] = field(default_factory=list)

    def as_dict(self) -> dict[str, Any]:
        return {
            "audits_created": self.audits_created,
            "audits_skipped": self.audits_skipped,
            "norms_created": len(self.norms_created),
            "norms_reused": self.norms_reused,
            "warnings": self.warnings,
        }


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    return " ".join(str(value).split())


_MONTHS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}


def parse_target_date(value: Any, warnings: list[str], label: str) -> datetime.date | None:
    """Parse a target date cell.

    Most cells are real dates. At least one ("Oct 27") is free text, which is
    why this tolerates a month-year form and maps it to the first of the month
    rather than dropping the date.
    """
    if isinstance(value, datetime.datetime):
        return value.date()
    text = _clean(value)
    if not text:
        return None
    m = re.match(r"^([A-Za-z]{3})[a-z]*\.?\s*'?(\d{2,4})$", text)
    if m and m.group(1).lower() in _MONTHS:
        year = int(m.group(2))
        year += 2000 if year < 100 else 0
        parsed = datetime.date(year, _MONTHS[m.group(1).lower()], 1)
        warnings.append(
            f"{label}: Ziel-Termin {text!r} ist Freitext — als {parsed.isoformat()} übernommen"
        )
        return parsed
    warnings.append(f"{label}: Ziel-Termin {text!r} nicht interpretierbar — leer gelassen")
    return None


def _split_clauses(regulation: str, raw: str) -> list[str]:
    """Split a reference cell into individual clause tokens."""
    text = raw.replace("QMH", " ")
    parts = re.split(r"[,/]" if regulation == "QMH" else r",", text)
    return [p.strip() for p in parts if p.strip()]


def parse_programme(path: str) -> tuple[list[ParsedAudit], list[str]]:
    """Read the workbook and return the audits plus any parsing warnings."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[SHEET]
    warnings: list[str] = []

    divisions = {r: _clean(ws.cell(row=r, column=1).value) for r in DIVISION_ROWS}
    audits: list[ParsedAudit] = []

    for _year_label, columns in YEAR_COLUMNS:
        for col in columns:
            title = _clean(ws[f"{col}{ROW['title']}"].value)
            if not title:
                continue
            nr = _clean(ws[f"{col}{ROW['nr']}"].value)
            label = f"Spalte {col} ({title})"

            planned_start = parse_target_date(
                ws[f"{col}{ROW['target']}"].value, warnings, label
            )
            # The audit number carries the calendar year of the target date, so
            # the repeating per-year numbering (1..6 three times) stays unique.
            year = planned_start.year if planned_start else 0
            audit_number = f"IA-{year}-{nr}" if year else f"IA-{nr}"

            is_process = _clean(ws[f"{col}{ROW['process']}"].value).upper().startswith("X")
            is_product = _clean(ws[f"{col}{ROW['product']}"].value).upper().startswith("X")
            categories = [c for c, on in (("prozess", is_process), ("produkt", is_product)) if on]
            if not categories:
                categories = ["prozess"]
                warnings.append(f"{label}: weder Part 1 noch Part 2 markiert — als Prozessaudit übernommen")

            scope = [divisions[r] for r in DIVISION_ROWS
                     if _clean(ws[f"{col}{r}"].value).upper().startswith("X")]
            if not scope:
                warnings.append(f"{label}: keine Bereiche markiert — Geltungsbereich bleibt leer")

            clauses: list[tuple[str, str, str]] = []
            reference_text: list[str] = []
            for row, regulation in REFERENCE_ROWS.items():
                raw = _clean(ws[f"{col}{row}"].value)
                if not raw:
                    continue
                reference_text.append(f"{regulation}: {raw}")
                for token in _split_clauses(regulation, raw):
                    clause = f"{CLAUSE_PREFIX.get(regulation, '')}{token}".strip()
                    clauses.append((regulation, REFERENCE_REVISION[regulation], clause))

            objective_parts = [_clean(ws[f"{col}{ROW['abstract']}"].value)]
            if is_product:
                objective_parts.append("Zusätzlich Produktaudit (Part 2).")
            if reference_text:
                objective_parts.append("Referenzen laut Auditprogramm — " + " | ".join(reference_text))

            audits.append(ParsedAudit(
                audit_number=audit_number,
                title=title,
                objective="\n\n".join(p for p in objective_parts if p),
                categories=categories,
                scope_label=", ".join(scope)[:255],
                lead_auditor=_clean(ws[f"{col}{ROW['auditor']}"].value) or None,
                planned_start=planned_start,
                actual_date_raw=_clean(ws[f"{col}{ROW['actual_date']}"].value),
                actual_auditor=_clean(ws[f"{col}{ROW['actual_auditor']}"].value),
                norm_clauses=clauses,
            ))

    return audits, warnings


async def _get_or_create_norm(
    db: AsyncSession,
    actor: CurrentUser,
    regulation: str,
    revision: str,
    clause: str,
    report: ImportReport,
    cache: dict[tuple[str, str, str], Any],
) -> Any:
    key = (regulation, revision, clause)
    if key in cache:
        return cache[key]

    found = await db.execute(
        sa.select(AuditNormReference).where(
            AuditNormReference.regulation == regulation,
            AuditNormReference.revision == revision,
            AuditNormReference.clause == clause,
        )
    )
    norm = found.scalar_one_or_none()
    if norm is not None:
        report.norms_reused += 1
        cache[key] = norm
        return norm

    # verified=False on purpose: these clause numbers come from the programme
    # spreadsheet and have not been checked against the regulation text.
    norm = AuditNormReference(
        regulation=regulation, revision=revision, clause=clause,
        short_text="", verified=False, active=True,
    )
    db.add(norm)
    await db.flush()
    record(
        db, actor=actor, entity_type="audit_norm_reference", entity_id=norm.id,
        action="create", field="clause", new_value=f"{regulation} {clause}",
        reason="Import Auditprogramm 2025-2027",
    )
    report.norms_created.append(f"{regulation} {clause}")
    cache[key] = norm
    return norm


async def import_programme(
    db: AsyncSession,
    actor: CurrentUser,
    parsed: list[ParsedAudit],
    warnings: list[str],
    *,
    template_id: Any | None = None,
    dry_run: bool = False,
) -> ImportReport:
    """Create the audits, their phases and their norm links. Idempotent."""
    report = ImportReport(warnings=list(warnings))
    norm_cache: dict[tuple[str, str, str], Any] = {}

    if template_id is None:
        found = await db.execute(
            sa.select(AuditPhaseTemplate)
            .where(AuditPhaseTemplate.active.is_(True))
            .order_by(AuditPhaseTemplate.created_at)
            .limit(1)
        )
        template = found.scalar_one_or_none()
    else:
        template = await db.get(AuditPhaseTemplate, template_id)
    if template is None:
        raise RuntimeError("no active phase template found — cannot instantiate phases")

    steps = (await db.execute(
        sa.select(AuditPhaseTemplateStep)
        .where(AuditPhaseTemplateStep.template_id == template.id)
        .order_by(AuditPhaseTemplateStep.position)
    )).scalars().all()

    for item in parsed:
        existing = await db.execute(
            sa.select(Audit.id).where(Audit.audit_number == item.audit_number)
        )
        if existing.scalar_one_or_none() is not None:
            report.audits_skipped.append(item.audit_number)
            continue

        audit = Audit(
            audit_number=item.audit_number,
            title=item.title,
            audit_type="intern",           # the whole programme is internal
            scope_label=item.scope_label,
            objective=item.objective,
            lead_auditor=item.lead_auditor,
            planned_start=item.planned_start,
            planned_end=item.planned_start,  # the sheet gives a single target date
            priority=2,
            status="geplant",              # never advanced by the importer
            template_id=template.id,
        )
        db.add(audit)
        await db.flush()

        for category in item.categories:
            db.add(AuditCategoryLink(audit_id=audit.id, category=category))

        for step in steps:
            phase = AuditPhase(
                audit_id=audit.id, position=step.position, title=step.title,
                description=step.description, mandatory=step.mandatory,
            )
            # Carry the recorded execution over as a comment rather than a
            # status, so a human still makes the completion decision.
            if step.title == EXECUTION_PHASE_TITLE and item.actual_date_raw:
                phase.comment = (
                    f"Laut Auditprogramm durchgeführt am {item.actual_date_raw}"
                    + (f" durch {item.actual_auditor}" if item.actual_auditor else "")
                )
            db.add(phase)

        seen: set[tuple[str, str, str]] = set()
        for regulation, revision, clause in item.norm_clauses:
            key = (regulation, revision, clause)
            if key in seen:
                continue
            seen.add(key)
            norm = await _get_or_create_norm(
                db, actor, regulation, revision, clause, report, norm_cache
            )
            db.add(AuditNormLink(audit_id=audit.id, norm_reference_id=norm.id))

        record(
            db, actor=actor, entity_type="audit", entity_id=audit.id,
            audit_id=audit.id, action="create", field="audit_number",
            new_value=audit.audit_number,
            reason="Import Auditprogramm 2025-2027 Version03",
        )
        report.audits_created.append(item.audit_number)

    if dry_run:
        await db.rollback()
    else:
        await db.commit()
    return report
