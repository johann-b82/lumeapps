"""Schulungsnachweis (intern) je Schulung als PDF — Formblatt 68 (v1.93).

Anders als die Schulungsübersicht (pro Person) ist dies ein Nachweis pro
Schulung: Titel/Datum/Dauer/Ziel, ein fester Bestätigungstext, eine nummerierte
Teilnehmerliste mit Unterschriftenspalte und der Trainer-Block.

Der Titel kommt aus dem Katalog; die übrigen Felder (Datum, Dauer, Trainer, die
Namen und Unterschriften) bleiben leer und werden bei der Schulung von Hand
ausgefüllt. Optional lassen sich Teilnehmer vorbelegen.

Aufbau wie die anderen Formblätter: openpyxl -> LibreOffice -> PDF.
"""
from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.page import PageMargins

from app.services.maintenance_pdf import convert_xlsx_to_pdf
from app.services.pdf_logo import LogoBild, bild_einsetzen

FORMBLATT = "Formblatt 68"
REVISIONS_INDEX = "A"
REVISIONS_STAND = "22.03.2022"

#: Feste Teilnahme-Bestätigung, wörtlich aus dem Formblatt.
BESTAETIGUNG = (
    "Ich habe an der oben genannten Schulung/Unterweisung teilgenommen. Die "
    "Thematik wurde angemessen verdeutlicht. Ich habe die Schulungsinhalte "
    "verstanden (und das dementsprechende Skript wurde ausgehändigt – falls "
    "vorhanden). Bei offenen Fragen wende ich mich an die entsprechend "
    "zuständige(n) Person(en)."
)

#: Anzahl Teilnehmerzeilen, wenn nichts vorbelegt ist — füllt ein A4-Blatt.
LEERE_ZEILEN = 12

_THIN = Side(style="thin", color="000000")
_RAHMEN = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _fuelle_blatt(
    ws,
    titel: str,
    teilnehmer: list[str],
    trainer: str,
    intern: bool,
    logo: LogoBild | None,
) -> None:
    ws.title = "Schulungsnachweis"
    # A = Nr.-Spalte der Teilnehmertabelle (schmal). Label-Zeilen oben/unten
    # legen A:B zusammen, damit die Beschriftungen ("Titel der …") vollständig
    # stehen; ihr Wert steht dann in C.
    for spalte, breite in zip("ABC", (6, 40, 42)):
        ws.column_dimensions[spalte].width = breite

    # Seitenkopf: Blatt x von y (nicht fest verdrahtet).
    ws.oddHeader.right.text = "&9Blatt &P von &N"
    ws.evenHeader.right.text = ws.oddHeader.right.text

    top = 1
    if logo is not None:
        bild_einsetzen(ws, logo, "A1")
        for r in (1, 2, 3):
            ws.row_dimensions[r].height = 20
        top = 4

    # Formblatt-Block rechts.
    for i, text in enumerate(
        [FORMBLATT, f"Revisions-Index: {REVISIONS_INDEX}", f"Revisions-Stand: {REVISIONS_STAND}"]
    ):
        z = ws.cell(row=top + i, column=3, value=text)
        z.font = Font(size=9)
        z.alignment = Alignment(horizontal="right", vertical="center")

    titelz = ws.cell(row=top, column=1, value="Schulungsnachweis (intern)")
    titelz.font = Font(size=16, bold=True)
    ws.merge_cells(start_row=top, start_column=1, end_row=top + 1, end_column=2)

    r = top + 3
    for label, wert in [
        ("Titel der Lehrveranstaltung:", titel),
        ("Datum der Lehrveranstaltung:", ""),
        ("Dauer der Lehrveranstaltung:", ""),
        ("Ziel/ Inhalt der Lehrveranstaltung:", ""),
    ]:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        z = ws.cell(row=r, column=3, value=wert)
        z.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        r += 1

    r += 1
    best = ws.cell(row=r, column=1, value=BESTAETIGUNG)
    best.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws.row_dimensions[r].height = 60
    r += 2

    # Teilnehmertabelle.
    kopf = ws.cell(row=r, column=1, value="Nr.")
    ws.cell(row=r, column=2, value="Teilnehmer (Name, Vorname)")
    ws.cell(row=r, column=3, value="Unterschrift der Teilnehmer")
    for c in range(1, 4):
        zelle = ws.cell(row=r, column=c)
        zelle.font = Font(size=9, bold=True)
        zelle.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        zelle.border = _RAHMEN
    _ = kopf
    r += 1

    namen = teilnehmer or [""] * LEERE_ZEILEN
    for i, name in enumerate(namen, start=1):
        ws.cell(row=r, column=1, value=i).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=2, value=name).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=r, column=3, value="")
        for c in range(1, 4):
            zelle = ws.cell(row=r, column=c)
            zelle.border = _RAHMEN
            zelle.font = Font(size=10)
        ws.row_dimensions[r].height = 24
        r += 1

    # Fußblock: Durchführung, Datum, Trainer. Labels über A:B, Wert in C.
    def _fussfeld(zeile: int, label: str, wert: str = "") -> None:
        ws.cell(row=zeile, column=1, value=label).font = Font(bold=True)
        ws.merge_cells(start_row=zeile, start_column=1, end_row=zeile, end_column=2)
        ws.cell(row=zeile, column=3, value=wert)

    r += 1
    haken_intern = "☒" if intern else "☐"
    haken_extern = "☐" if intern else "☒"
    _fussfeld(r, "Durchführung:", f"{haken_intern} intern     {haken_extern} extern")
    r += 2
    _fussfeld(r, "Datum:")
    r += 2
    _fussfeld(r, "Name des Trainers:", trainer)
    r += 2
    _fussfeld(r, "Unterschrift des Trainers:", "…………………………………………")

    ws.print_area = f"A1:C{r + 1}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.6, right=0.5, top=0.5, bottom=0.5, header=0.3)


async def erzeuge_schulungsprotokoll_pdf(
    titel: str,
    teilnehmer: list[str] | None = None,
    trainer: str = "",
    intern: bool = True,
    logo: LogoBild | None = None,
) -> bytes:
    wb = Workbook()
    _fuelle_blatt(wb.active, titel, teilnehmer or [], trainer, intern, logo)
    puffer = BytesIO()
    wb.save(puffer)
    return await convert_xlsx_to_pdf(puffer.getvalue())


def dateiname(titel: str, stand: date) -> str:
    teil = "_".join(titel.split())[:60] or "Schulung"
    return f"{stand:%Y.%m.%d}_{teil}_Schulungsnachweis"
