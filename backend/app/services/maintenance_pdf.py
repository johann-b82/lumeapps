"""Printable "Wartungsnachweis" sheet generation for a machine (v1.82).

Builds an .xlsx with openpyxl, then converts it to PDF via LibreOffice headless
(``soffice --convert-to pdf``) — the same LibreOffice that the ATR module already
relies on in the API image. No UNO scripting is needed here: everything (machine
header, KW/day grid, legend) lives in sheet cells.

Two sheet types, kept deliberately lean so nothing is overloaded:
  * "Periodisch" — a half-year sheet, KW columns (1..26 or 27..52) across the
    top, one row per non-daily task grouped by interval. The employee stamps /
    signs the cell of the calendar week in which the maintenance was performed.
  * "Täglich"    — only added when daily tasks exist: a month sheet with day
    columns (1..31) and a blank "Monat/Jahr" field to fill in by hand.
"""
from __future__ import annotations

import asyncio
import shutil
import uuid as _uuid
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

# Interval ordering + German labels used to group rows on the periodic sheet.
_PERIODIC_ORDER = ["weekly", "monthly", "quarterly", "interval_weeks"]
_INTERVAL_LABELS = {
    "daily": "Täglich",
    "weekly": "Wöchentlich",
    "monthly": "Monatlich",
    "quarterly": "Quartalsweise",
    "interval_weeks": "Alle N Wochen",
}

_THIN = Side(style="thin", color="9CA3AF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill("solid", fgColor="E5E7EB")
_GROUP_FILL = PatternFill("solid", fgColor="F3F4F6")
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_LEFT_NOWRAP = Alignment(horizontal="left", vertical="center", wrap_text=False)


def _interval_label(task) -> str:
    if task.interval_type == "interval_weeks" and task.interval_weeks:
        return f"Alle {task.interval_weeks} Wochen"
    return _INTERVAL_LABELS.get(task.interval_type, task.interval_type)


def _machine_header(ws, machine, subtitle: str, ncols: int) -> int:
    """Write the title + machine-info block. Returns the next free row (1-based)."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1, "Wartungsnachweis")
    c.font = Font(bold=True, size=16)
    c.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    s = ws.cell(2, 1, subtitle)
    s.font = Font(bold=True, size=11, color="374151")
    s.alignment = Alignment(horizontal="left", vertical="center")

    info = [
        ("Maschine", machine.name or "—"),
        ("Inventar-Nr.", machine.inventory_no or "—"),
        ("Standort", machine.location or "—"),
        ("Verantwortlich", machine.responsible or "—"),
    ]
    # Value cells are merged across a readable span so they never wrap into the
    # narrow KW / day grid columns. Labels sit in the (wide) first column.
    val_end = min(ncols, 8)
    for i, (label, value) in enumerate(info):
        row = 3 + i
        lc = ws.cell(row, 1, f"{label}:")
        lc.font = Font(bold=True, size=10)
        lc.alignment = _LEFT_NOWRAP
        if val_end > 2:
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=val_end)
        vc = ws.cell(row, 2, str(value))
        vc.font = Font(size=10)
        vc.alignment = _LEFT_NOWRAP
    return 3 + len(info) + 1  # blank spacer row after the block


def _grid_sheet(
    ws,
    machine,
    subtitle: str,
    first_col_header: str,
    second_col_header: str | None,
    col_labels: list[str],
    rows: list[tuple[str, str]],
) -> None:
    """Render one grid sheet.

    ``rows`` are ``(group_or_interval_label, task_title)`` tuples. When
    ``second_col_header`` is None there is a single label column (the daily
    sheet); otherwise the first column carries the interval group label.
    """
    label_cols = 2 if second_col_header else 1
    ncols = label_cols + len(col_labels)

    header_row = _machine_header(ws, machine, subtitle, ncols)

    # column header row
    ws.cell(header_row, 1, first_col_header)
    if second_col_header:
        ws.cell(header_row, 2, second_col_header)
    for j, lbl in enumerate(col_labels):
        ws.cell(header_row, label_cols + 1 + j, lbl)
    for col in range(1, ncols + 1):
        hc = ws.cell(header_row, col)
        hc.font = Font(bold=True, size=9)
        hc.alignment = _CENTER
        hc.fill = _HEADER_FILL
        hc.border = _BORDER

    # task rows
    r = header_row + 1
    for group_label, title in rows:
        if second_col_header:
            gc = ws.cell(r, 1, group_label)
            gc.font = Font(size=9)
            gc.alignment = _LEFT
            gc.fill = _GROUP_FILL
            gc.border = _BORDER
            tc = ws.cell(r, 2, title)
        else:
            tc = ws.cell(r, 1, title)
        tc.font = Font(size=9)
        tc.alignment = _LEFT
        tc.border = _BORDER
        for j in range(len(col_labels)):
            gcell = ws.cell(r, label_cols + 1 + j)
            gcell.border = _BORDER
        ws.row_dimensions[r].height = 26
        r += 1

    if not rows:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        e = ws.cell(r, 1, "Keine Aufgaben für diesen Bereich definiert.")
        e.font = Font(italic=True, size=9, color="6B7280")
        r += 1

    # legend
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    lg = ws.cell(
        r,
        1,
        "Bestätigung der durchgeführten Wartung durch Unterschrift / Stempel in "
        "der jeweiligen Spalte.",
    )
    lg.font = Font(italic=True, size=9, color="374151")
    lg.alignment = _LEFT

    # column widths
    ws.column_dimensions[get_column_letter(1)].width = 20 if second_col_header else 40
    if second_col_header:
        ws.column_dimensions[get_column_letter(2)].width = 40
    for j in range(len(col_labels)):
        ws.column_dimensions[get_column_letter(label_cols + 1 + j)].width = 4.2

    # print setup — landscape, fit to one page wide
    ws.print_area = f"A1:{get_column_letter(ncols)}{r}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)


def build_maintenance_xlsx(machine, tasks, year: int, half: int) -> bytes:
    """Build the .xlsx workbook (periodic sheet + optional daily sheet)."""
    wb = Workbook()

    # ── periodic (KW) sheet ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Periodisch"
    if half == 2:
        weeks = list(range(27, 53))
    else:
        weeks = list(range(1, 27))
    subtitle = f"Jahr {year} · KW {weeks[0]:02d}–{weeks[-1]:02d}"

    periodic = [t for t in tasks if t.interval_type in _PERIODIC_ORDER]
    periodic.sort(key=lambda t: (_PERIODIC_ORDER.index(t.interval_type), t.created_at))
    prows = [(_interval_label(t), t.title) for t in periodic]
    _grid_sheet(
        ws,
        machine,
        subtitle,
        first_col_header="Intervall",
        second_col_header="Wartungsaufgabe",
        col_labels=[f"KW {w:02d}" for w in weeks],
        rows=prows,
    )

    # ── daily sheet (only when daily tasks exist) ────────────────────────
    daily = [t for t in tasks if t.interval_type == "daily"]
    if daily:
        wd = wb.create_sheet("Täglich")
        drows = [("", t.title) for t in daily]  # single label column
        _grid_sheet(
            wd,
            machine,
            f"Tägliche Wartung · Monat / Jahr: ____________ / {year}",
            first_col_header="Wartungsaufgabe",
            second_col_header=None,
            col_labels=[str(d) for d in range(1, 32)],
            rows=[(g, t) for g, t in drows],
        )

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# Serialize LibreOffice across the single-worker api container (mirror ATR /
# signage_pptx). A unique UserInstallation profile per call avoids the shared
# profile lock even if another module's LibreOffice runs concurrently.
_LO_SEMAPHORE = asyncio.Semaphore(1)
_LO_TIMEOUT_S = 60


async def convert_xlsx_to_pdf(xlsx_bytes: bytes) -> bytes:
    async with _LO_SEMAPHORE:
        tempdir = Path(f"/tmp/maint_{_uuid.uuid4()}")
        try:
            tempdir.mkdir(parents=True, exist_ok=True)
            src = tempdir / "sheet.xlsx"
            src.write_bytes(xlsx_bytes)
            profile = tempdir / "profile"
            proc = await asyncio.create_subprocess_exec(
                "soffice", "--headless", "--invisible", "--nodefault",
                "--norestore", "--nologo", "--nofirststartwizard",
                f"-env:UserInstallation=file://{profile}",
                "--convert-to", "pdf:calc_pdf_Export",
                "--outdir", str(tempdir), str(src),
                stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE,
            )
            try:
                _, err = await asyncio.wait_for(
                    proc.communicate(), timeout=_LO_TIMEOUT_S
                )
            except asyncio.TimeoutError as exc:
                try:
                    proc.kill()
                except ProcessLookupError:
                    pass
                raise RuntimeError("xlsx->pdf conversion timed out") from exc
            out = tempdir / "sheet.pdf"
            if proc.returncode != 0 or not out.exists():
                raise RuntimeError(
                    f"soffice pdf export failed: {err.decode('utf-8', 'replace')[-500:]}"
                )
            return out.read_bytes()
        finally:
            shutil.rmtree(tempdir, ignore_errors=True)


async def generate_maintenance_pdf(machine, tasks, year: int, half: int) -> bytes:
    xlsx = build_maintenance_xlsx(machine, tasks, year, half)
    return await convert_xlsx_to_pdf(xlsx)
