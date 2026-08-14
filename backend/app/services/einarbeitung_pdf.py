"""Einarbeitungsbogen je Person als PDF (v1.92).

Bildet die Vorlage "Einarbeitungsplan" nach: ACM-Kopfzeilentabelle (Fbl.-Nummer,
Titel, Revision, Logo), Kopf mit Name/Stelle/Beginn/Einarbeitungszeitraum, feste
Einleitung (Regeln + Ziele), die Inhalts-Tabelle (Abteilung · Ansprechpartner ·
Inhalt · Wann? · Erledigt/Unterschrift), der Laufweg als kompakte Tabelle und
zuletzt — am unteren Seitenrand — die Freigabe-Fußzeile (Erstellt/Geprüft/
Freigegeben).

Aufbau wie die Schulungsübersicht: openpyxl schreibt eine .xlsx, LibreOffice
headless macht daraus ein PDF. Damit die Fußzeile verlässlich am Blattfuß sitzt
und Zeilenhöhen exakt bleiben, wird ohne Seitenskalierung gedruckt (Scale 100 %,
Spaltenbreiten passen auf A4).
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date, timedelta
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

from app.services.maintenance_pdf import convert_xlsx_to_pdf
from app.services.pdf_logo import LogoBild, bild_einsetzen

FORMBLATT = "Einarbeitungsplan"

#: Kopfzeilentabelle (ACM-Standard). Bei einer neuen Revision hier anpassen.
KOPF_FBL = "Fbl. 28 Einarbeitungsplan Rev. C vom 14.08.2026"
KOPF_TITEL = "Einarbeitungsplan"
REV_INDEX = "C"
REV_STAND = "14.08.2026"

#: Änderungsbeschreibung der aktuellen Revision (Revisionshistorie im Freigabe-Dok.).
REV_BESCHREIBUNG = (
    "Anpassung des Formulars für die automatische Erstellung. Keine inhaltlichen "
    "Änderungen; ausschließlich optische und technische Anpassungen."
)
REV_ERSTELLER = "M. Brose (QM/CMM)"

#: Dauer der Einarbeitung ab Tätigkeitsbeginn (für das vorausgefüllte "bis:").
EINARBEITUNG_DAUER = timedelta(days=28)  # 4 Wochen

#: Freigabe-Fußzeile am Formblatt (Platzhalter, wird je Ausdruck ausgefüllt).
FORM_ROLLEN = [
    ("Erstellt durch:", "[Name] (QM/CMM)"),
    ("Geprüft durch:", "[Name] (TBL)"),
    ("Freigegeben durch:", "[Name] (QM/CMM)"),
]
#: Freigabe-Fußzeile auf dem eigenständigen Freigabe-Dokument (mit Klarnamen).
FREIGABE_ROLLEN = [
    ("Erstellt durch:", "M. Brose (QM/CMM)"),
    ("Geprüft durch:", "F. Gomes (TBL)"),
    ("Freigegeben durch:", "M. Brose (QM/CMM)"),
]

#: Spalten wie in der Vorlage: A ist eine schmale Gutter-Spalte, dann B=Abteilung,
#: C=Ansprechpartner, D:F=Inhalt, G=Wann?, H=Erledigt/Unterschrift.
SP_ABT, SP_PART, SP_INHALT, SP_WANN, SP_ERLEDIGT = 2, 3, 4, 7, 8

#: Drei Zellen der Freigabe-Fußzeile (Start-/End-Spalte), passend zu B..H.
_FREIGABE_SPALTEN = [(2, 3), (4, 5), (6, 8)]

#: Spaltenbreiten des Blatts (A..H). Summe so gewählt, dass A1:H auf A4 hochkant
#: bei 100 % passt (keine Druckskalierung → exakte Zeilenhöhen).
_SPALTENBREITEN = (3, 13, 17, 13, 13, 13, 8, 14)

#: Zeichen-je-Zeile für den Umbruch. Bezug ist der 9-pt-Tabellentext, der in
#: einer Spaltenbreite deutlich mehr Zeichen fasst als eine Breiteneinheit
#: (die auf der 11-pt-Standardschrift beruht) — daher die höheren Werte.
_CH_ABT, _CH_PART, _CH_INHALT = 14, 19, 46

#: Zeilenhöhe je Textzeile (9 pt) inkl. kleinem Innenabstand.
_ZEILE_PT, _ZEILE_PAD = 12.0, 5.0

#: Nutzhöhe einer A4-Druckseite in openpyxl-Zeilenhöhen-Einheiten. Empirisch
#: kalibriert (39 Zeilen à 20 ≈ 780 passen), mit Sicherheitsabstand darunter,
#: da die summierten Höhen die tatsächliche Darstellung leicht unterschätzen —
#: so rutscht die Fußzeile sicher an den Blattfuß, ohne über die Seitenkante zu
#: kippen. benutzt und Kapazität sind in derselben Einheit — keine Umrechnung.
_SEITE_KAP_PT = 748.0
_DEFAULT_ROW_PT = 15.0

_REGELN = [
    "Jeder neue Mitarbeiter wird anhand eines Einarbeitungsplans systematisch eingelernt.",
    "Die einzelnen Einarbeitungsschritte werden vom Mitarbeiter dokumentiert.",
    "Innerhalb der ersten 4 Wochen findet ein Feedbackgespräch statt.",
]
_ZIELE = (
    "Ziel der Einarbeitung ist die gezielte Unterstützung der Abteilung durch den "
    "Aufbau der notwendigen Kenntnisse und Fähigkeiten in den relevanten Prozessen."
)

_THIN = Side(style="thin", color="000000")
_RAHMEN = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_UNTERLINIE = Border(bottom=_THIN)
_LINKS_OBEN = Alignment(horizontal="left", vertical="top", wrap_text=True)
_LINKS = Alignment(horizontal="left", vertical="center")
_ZENTRIERT = Alignment(horizontal="center", vertical="center", wrap_text=True)


@dataclass
class EinarbeitungZeile:
    abteilung: str
    inhalt: str
    ansprechpartner: str = ""


def _rahmen_bereich(ws, r1: int, c1: int, r2: int, c2: int) -> None:
    """Dünnen Rahmen um jede Zelle eines Rechtecks legen (auch bei Merges)."""
    for rr in range(r1, r2 + 1):
        for cc in range(c1, c2 + 1):
            ws.cell(row=rr, column=cc).border = _RAHMEN


def _linie(ws, r: int, c1: int, c2: int) -> None:
    """Beschreibbare Linie (Unterstrich) über einen Spaltenbereich."""
    if c2 > c1:
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    for c in range(c1, c2 + 1):
        ws.cell(row=r, column=c).border = _UNTERLINIE


def _spalten_setzen(ws) -> None:
    for spalte, breite in zip("ABCDEFGH", _SPALTENBREITEN):
        ws.column_dimensions[spalte].width = breite


def _umbruch_zeilen(text: str, breite: int) -> int:
    """Anzahl Zeilen, die ``text`` bei ``breite`` Zeichen umbrochen belegt."""
    text = (text or "").strip()
    if not text:
        return 1
    gesamt = 0
    for absatz in text.split("\n"):
        woerter = absatz.split()
        if not woerter:
            gesamt += 1
            continue
        zeile = ""
        n = 0
        for w in woerter:
            probe = w if not zeile else f"{zeile} {w}"
            if len(probe) <= breite:
                zeile = probe
            else:
                n += 1
                zeile = w
        gesamt += n + 1
    return max(1, gesamt)


def _kopfzeilen_tabelle(ws, logo: LogoBild | None) -> int:
    """ACM-Kopfzeilentabelle in Zeilen 1–3 (Fbl./Titel · Revision · Logo).

    Gibt die erste freie Inhaltszeile zurück (Zeile 4 bleibt als Abstand frei).
    """
    for r in (1, 2, 3):
        ws.row_dimensions[r].height = 16

    # Linker Block: Fbl.-Zeile über Titel.
    ws.merge_cells("B1:E1")
    fbl = ws.cell(row=1, column=2, value=KOPF_FBL)
    fbl.font = Font(size=10, color="595959")
    fbl.alignment = _ZENTRIERT
    ws.merge_cells("B2:E3")
    titel = ws.cell(row=2, column=2, value=KOPF_TITEL)
    titel.font = Font(size=16, bold=True)
    titel.alignment = _ZENTRIERT

    # Mittlerer Block: Revisionsangaben (Seitenzahl steht dynamisch in der Fußzeile).
    ws.merge_cells("F1:G3")
    z = ws.cell(
        row=1,
        column=6,
        value=f"Revisions-Index: {REV_INDEX}\nRevisions-Stand: {REV_STAND}",
    )
    z.font = Font(size=8)
    z.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Rechter Block: Logo (auf Kopfzeilenhöhe verkleinert).
    ws.merge_cells("H1:H3")
    if logo is not None:
        breite = 90
        kopf_logo = LogoBild(
            daten=logo.daten,
            breite=breite,
            hoehe=max(1, round(breite * logo.hoehe / logo.breite)),
        )
        bild_einsetzen(ws, kopf_logo, "H1")

    _rahmen_bereich(ws, 1, 2, 3, 8)
    return 5


def _kopf(ws, name: str, stelle: str, beginn: date | None, logo: LogoBild | None = None) -> int:
    r = _kopfzeilen_tabelle(ws, logo)

    zeilen = [
        ("Name des Mitarbeiters:", name),
        ("Stellenbezeichnung:", stelle or "—"),
        ("Beginn der Tätigkeit:", beginn.strftime("%d.%m.%Y") if beginn else "—"),
    ]
    for label, wert in zeilen:
        ws.cell(row=r, column=2, value=label).font = Font(bold=True)
        z = ws.cell(row=r, column=4, value=wert)
        z.alignment = _LINKS
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)
        ws.row_dimensions[r].height = 16
        r += 1

    # Einarbeitungszeitraum vorausgefüllt: von = Beginn, bis = Beginn + 4 Wochen.
    von = beginn.strftime("%d.%m.%Y") if beginn else ""
    bis = (beginn + EINARBEITUNG_DAUER).strftime("%d.%m.%Y") if beginn else ""
    ws.cell(row=r, column=2, value="Einarbeitungszeit von:").font = Font(bold=True)
    ws.cell(row=r, column=4, value=von).alignment = _LINKS
    ws.cell(row=r, column=5, value="bis:").font = Font(bold=True)
    ws.cell(row=r, column=6, value=bis).alignment = _LINKS
    ws.row_dimensions[r].height = 16
    r += 2

    # Feedbackgespräch: Bezeichnung + klar beschriftete Felder zum Ausfüllen.
    ws.cell(row=r, column=2, value="Feedbackgespräch nach spätestens 4 Wochen:").font = Font(bold=True)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    ws.row_dimensions[r].height = 16
    r += 1
    ws.cell(row=r, column=2, value="Datum:").font = Font(bold=True)
    _linie(ws, r, 3, 5)
    ws.row_dimensions[r].height = 20
    r += 2
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    ws.cell(row=r, column=2, value="Unterschrift Vorgesetzter:").font = Font(bold=True)
    _linie(ws, r, 5, 8)
    ws.row_dimensions[r].height = 20
    return r + 1


def _einleitung(ws, r: int) -> int:
    ws.cell(row=r, column=2, value="Wichtige Einarbeitungsregeln:").font = Font(bold=True)
    ws.row_dimensions[r].height = 15
    r += 1
    for regel in _REGELN:
        ws.cell(row=r, column=2, value=f"• {regel}")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        ws.row_dimensions[r].height = 15
        r += 1
    r += 1
    ws.cell(row=r, column=2, value="Ziele der Einarbeitung:").font = Font(bold=True)
    ws.row_dimensions[r].height = 15
    r += 1
    ziel = ws.cell(row=r, column=2, value=_ZIELE)
    ziel.alignment = _LINKS_OBEN
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    ws.row_dimensions[r].height = 28
    return r + 2


def _tabelle(ws, r: int, zeilen: list[EinarbeitungZeile]) -> int:
    kopf = {
        SP_ABT: "Abteilung",
        SP_PART: "Ansprechpartner",
        SP_INHALT: "Inhalte der Einarbeitung",
        SP_WANN: "Wann?",
        SP_ERLEDIGT: "Erledigt/\nUnterschrift",
    }
    for c, text in kopf.items():
        zelle = ws.cell(row=r, column=c, value=text)
        zelle.font = Font(size=9, bold=True)
        zelle.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        zelle.border = _RAHMEN
    ws.merge_cells(start_row=r, start_column=SP_INHALT, end_row=r, end_column=SP_WANN - 1)
    ws.row_dimensions[r].height = 26
    r += 1

    for z in zeilen:
        ws.cell(row=r, column=SP_ABT, value=z.abteilung)
        ws.cell(row=r, column=SP_PART, value=z.ansprechpartner)
        ws.cell(row=r, column=SP_INHALT, value=z.inhalt)
        ws.merge_cells(
            start_row=r, start_column=SP_INHALT, end_row=r, end_column=SP_WANN - 1
        )
        for c in range(SP_ABT, SP_ERLEDIGT + 1):
            zelle = ws.cell(row=r, column=c)
            zelle.border = _RAHMEN
            zelle.font = Font(size=9)
            zelle.alignment = _LINKS_OBEN
        # Höhe an den tatsächlichen (umbrochenen) Inhalt anpassen, damit nichts
        # abgeschnitten wird — die höchste der drei Textspalten bestimmt sie.
        zeilen_anzahl = max(
            _umbruch_zeilen(z.abteilung, _CH_ABT),
            _umbruch_zeilen(z.ansprechpartner, _CH_PART),
            _umbruch_zeilen(z.inhalt, _CH_INHALT),
        )
        ws.row_dimensions[r].height = max(22.0, zeilen_anzahl * _ZEILE_PT + _ZEILE_PAD)
        r += 1
    return r


def _schulungsbedarf(ws, r: int) -> int:
    r += 1
    ws.cell(row=r, column=2, value="Weiterer Schulungsbedarf notwendig?").font = Font(bold=True)
    ws.cell(row=r, column=6, value="☐ ja    ☐ nein")
    ws.row_dimensions[r].height = 15
    r += 1
    ws.cell(row=r, column=2, value="Falls ja, bitte Schulungsbedarf erläutern:")
    ws.row_dimensions[r].height = 15
    r += 1
    _linie(ws, r, 2, 8)
    ws.row_dimensions[r].height = 18
    return r + 1


def _freigabe_fuss(ws, r: int, rollen: list[tuple[str, str]], mit_unterschrift: bool) -> int:
    """Freigabe-Fußzeile (Erstellt/Geprüft/Freigegeben) als drei Zellen.

    ``mit_unterschrift`` ergänzt je Zelle die Felder ``Datum:`` und
    ``Unterschrift:`` (für das eigenständige Freigabe-Dokument).
    """
    zeilen_je_spalte = []
    for rolle, name in rollen:
        lines = [(rolle, True), (name, False)]
        if mit_unterschrift:
            lines += [("Datum:", False), ("Unterschrift:", False)]
        zeilen_je_spalte.append(lines)

    n = len(zeilen_je_spalte[0])
    for i in range(n):
        for (c1, c2), lines in zip(_FREIGABE_SPALTEN, zeilen_je_spalte):
            text, bold = lines[i]
            zelle = ws.cell(row=r + i, column=c1, value=text)
            zelle.font = Font(size=9, bold=bold)
            zelle.alignment = _LINKS
            if c2 > c1:
                ws.merge_cells(start_row=r + i, start_column=c1, end_row=r + i, end_column=c2)
        ws.row_dimensions[r + i].height = 22 if i >= 2 else 16

    _rahmen_bereich(ws, r, 2, r + n - 1, 8)
    return r + n


def _revisionshistorie(ws, r: int) -> int:
    """Revisionshistorie als Tabelle (Revision · Datum · Beschreibung · durch)."""
    ws.cell(row=r, column=2, value="Revisionshistorie").font = Font(size=11, bold=True)
    ws.row_dimensions[r].height = 16
    r += 1

    spalten = [
        ((2, 2), "Revision"),
        ((3, 3), "Datum"),
        ((4, 6), "Beschreibung der Änderung"),
        ((7, 8), "Erstellt / geändert durch"),
    ]
    for (c1, c2), text in spalten:
        z = ws.cell(row=r, column=c1, value=text)
        z.font = Font(size=9, bold=True)
        z.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if c2 > c1:
            ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    _rahmen_bereich(ws, r, 2, r, 8)
    ws.row_dimensions[r].height = 26
    r += 1

    werte = [
        ((2, 2), REV_INDEX),
        ((3, 3), REV_STAND),
        ((4, 6), REV_BESCHREIBUNG),
        ((7, 8), REV_ERSTELLER),
    ]
    for (c1, c2), text in werte:
        z = ws.cell(row=r, column=c1, value=text)
        z.font = Font(size=9)
        z.alignment = _LINKS_OBEN
        if c2 > c1:
            ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    _rahmen_bereich(ws, r, 2, r, 8)
    zeilen_anzahl = max(
        _umbruch_zeilen(REV_BESCHREIBUNG, _CH_INHALT),
        _umbruch_zeilen("Erstellt / geändert durch", 21),
    )
    ws.row_dimensions[r].height = max(28.0, zeilen_anzahl * _ZEILE_PT + _ZEILE_PAD)
    return r + 1


def _freigabe_hoehe(mit_unterschrift: bool) -> float:
    return (16.0 + 16.0) + (22.0 + 22.0 if mit_unterschrift else 0.0)


def _fuss_ans_seitenende(ws, naechste_zeile: int, fuss_hoehe_pt: float) -> int:
    """Fußzeile an den Blattfuß schieben, ohne sie über die Seitenkante zu trennen.

    Passt der Inhalt auf eine Seite, werden Leerzeilen bis zum Seitenfuß ergänzt.
    Reicht der Platz für die Fußzeile nicht mehr, wird die Seite gefüllt, ein
    Umbruch gesetzt und die Fußzeile geschlossen an den Fuß der Folgeseite
    gelegt. Bei ohnehin mehrseitigem Inhalt folgt die Fußzeile direkt. Voraus
    werden alle noch unbesetzten Zeilenhöhen fixiert, damit die Höhensumme der
    Darstellung entspricht.
    """
    for i in range(1, naechste_zeile):
        if ws.row_dimensions[i].height is None:
            ws.row_dimensions[i].height = _DEFAULT_ROW_PT
    benutzt = sum(ws.row_dimensions[i].height for i in range(1, naechste_zeile))

    r = naechste_zeile

    def _fuellen(anzahl: int) -> None:
        nonlocal r
        for _ in range(max(0, anzahl)):
            ws.row_dimensions[r].height = _DEFAULT_ROW_PT
            r += 1

    if benutzt >= _SEITE_KAP_PT:
        # Inhalt füllt schon mehr als eine Seite — Fußzeile folgt unmittelbar.
        return r

    rest = _SEITE_KAP_PT - benutzt
    if fuss_hoehe_pt <= rest:
        _fuellen(int((rest - fuss_hoehe_pt) // _DEFAULT_ROW_PT))
    else:
        # Fußzeile passt nicht mehr → Seite füllen, umbrechen, unten auf Seite 2.
        _fuellen(int(rest // _DEFAULT_ROW_PT))
        ws.row_breaks.append(Break(id=r - 1))
        _fuellen(int((_SEITE_KAP_PT - fuss_hoehe_pt) // _DEFAULT_ROW_PT))
    return r


def _seiteneinrichtung(ws, letzte_zeile: int, wiederhol_kopf: int | None = None) -> None:
    ws.print_area = f"A1:H{letzte_zeile + 1}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.scale = 100  # keine Skalierung → exakte Zeilenhöhen
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=False)
    ws.page_margins = PageMargins(
        left=0.5, right=0.4, top=0.5, bottom=0.5, header=0.3, footer=0.25
    )
    # Dynamische Seitenzählung (LibreOffice-Fußzeile, auf jeder Seite korrekt).
    ws.oddFooter.right.text = "Blatt &[Page] von &[Pages]"
    ws.oddFooter.right.size = 8
    if wiederhol_kopf is not None:
        ws.print_title_rows = f"{wiederhol_kopf}:{wiederhol_kopf}"


def fuelle_blatt(
    ws,
    name: str,
    stelle: str,
    beginn: date | None,
    zeilen: list[EinarbeitungZeile],
    logo: LogoBild | None = None,
) -> None:
    """Einarbeitungsplan in ein vorhandenes Arbeitsblatt schreiben.

    Herausgezogen aus ``baue_xlsx``, damit das Blatt auch als erste Seite eines
    kombinierten Onboarding-Pakets (mit der Schulungsübersicht) dienen kann.
    """
    ws.title = "Einarbeitungsplan"
    _spalten_setzen(ws)

    r = _kopf(ws, name, stelle, beginn, logo)
    r = _einleitung(ws, r)
    tab_start = r
    r = _tabelle(ws, r, zeilen)
    r = _schulungsbedarf(ws, r)
    r += 1
    r = _fuss_ans_seitenende(ws, r, _freigabe_hoehe(mit_unterschrift=False))
    r = _freigabe_fuss(ws, r, FORM_ROLLEN, mit_unterschrift=False)

    # Läuft die Liste auf eine zweite Seite, wiederholt sich der Tabellenkopf.
    _seiteneinrichtung(ws, r, wiederhol_kopf=tab_start)


def baue_xlsx(
    name: str,
    stelle: str,
    beginn: date | None,
    zeilen: list[EinarbeitungZeile],
    logo: LogoBild | None = None,
) -> bytes:
    wb = Workbook()
    fuelle_blatt(wb.active, name, stelle, beginn, zeilen, logo)
    puffer = BytesIO()
    wb.save(puffer)
    return puffer.getvalue()


async def erzeuge_einarbeitung_pdf(
    name: str,
    stelle: str,
    beginn: date | None,
    zeilen: list[EinarbeitungZeile],
    logo: LogoBild | None = None,
) -> bytes:
    return await convert_xlsx_to_pdf(baue_xlsx(name, stelle, beginn, zeilen, logo))


def baue_freigabe_xlsx(logo: LogoBild | None = None) -> bytes:
    """Eigenständiges Freigabe-Dokument für die Formblatt-Revision.

    Gleiche ACM-Kopfzeile wie das Formblatt, ein kurzer Freigabesatz und die
    Freigabe-Fußzeile (mit Klarnamen sowie Datum-/Unterschrift-Feldern) am
    unteren Seitenrand.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Freigabe"
    _spalten_setzen(ws)

    r = _kopfzeilen_tabelle(ws, logo)

    ws.cell(row=r, column=2, value="Freigabe-Dokument").font = Font(size=14, bold=True)
    ws.row_dimensions[r].height = 20
    r += 2

    for label, wert in (
        ("Formblatt:", "Fbl. 28 Einarbeitungsplan"),
        ("Revision:", REV_INDEX),
        ("Revisions-Stand:", REV_STAND),
    ):
        ws.cell(row=r, column=2, value=label).font = Font(bold=True)
        wz = ws.cell(row=r, column=4, value=wert)
        wz.alignment = _LINKS
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
        ws.row_dimensions[r].height = 16
        r += 1
    r += 1

    satz = ws.cell(
        row=r,
        column=2,
        value=(
            "Hiermit wird das Formblatt Fbl. 28 „Einarbeitungsplan“ in der Revision "
            f"{REV_INDEX} mit Stand {REV_STAND} zur Anwendung freigegeben."
        ),
    )
    satz.alignment = _LINKS_OBEN
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    ws.row_dimensions[r].height = 30
    r += 2

    r = _revisionshistorie(ws, r)

    r = _fuss_ans_seitenende(ws, r, _freigabe_hoehe(mit_unterschrift=True))
    r = _freigabe_fuss(ws, r, FREIGABE_ROLLEN, mit_unterschrift=True)

    _seiteneinrichtung(ws, r)
    puffer = BytesIO()
    wb.save(puffer)
    return puffer.getvalue()


async def erzeuge_freigabe_pdf(logo: LogoBild | None = None) -> bytes:
    return await convert_xlsx_to_pdf(baue_freigabe_xlsx(logo))


def dateiname(name: str, stand: date) -> str:
    teil = "_".join(name.split()) or "Unbekannt"
    return f"{stand:%Y.%m.%d}_{teil}_Einarbeitungsplan"
